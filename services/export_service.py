"""
Export service — Generate factory order Excel files.

Converts Order Builder selections into the factory's expected format.
Also generates BL allocation reports for customs safety spreading.
"""

import re
from datetime import date
from decimal import Decimal
from io import BytesIO
from math import ceil
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import structlog

from models.bl_allocation import BLAllocationReport
from models.order_builder import (
    OrderBuilderResponse,
    OrderBuilderProduct,
    AddToProductionItem,
    FactoryRequestItem,
)

logger = structlog.get_logger(__name__)

# Factory constants (actual factory pallet dimensions)
M2_PER_PALLET_FACTORY = Decimal("134.4")
PALLETS_PER_CONTAINER = 14

# Spanish month names
MONTHS_ES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}


def normalize_sku_for_factory(sku: str) -> str:
    """
    Convert system SKU to factory format.

    'ALMENDRO BEIGE BTE (T) 51X51-1' -> 'ALMENDRO BEIGE'
    'CEIBA GRIS CLARO BTE' -> 'CEIBA GRIS CLARO'
    'NOGAL CAFE BTE' -> 'NOGAL CAFE'
    """
    if not sku:
        return ""

    # Remove common suffixes in order
    sku = re.sub(r"\s*\(T\)", "", sku)  # Remove (T)
    sku = re.sub(r"\s*BTE\b", "", sku, flags=re.IGNORECASE)  # Remove BTE
    sku = re.sub(r"\s*\d+X\d+.*$", "", sku)  # Remove 51X51-1 and anything after
    sku = re.sub(r"\s*-\d+$", "", sku)  # Remove trailing -1, -2, etc.

    return sku.strip()


def calculate_m2(pallets: int) -> Decimal:
    """Calculate m2 from pallets using factory constant."""
    return Decimal(pallets) * M2_PER_PALLET_FACTORY


def calculate_containers(total_pallets: int) -> int:
    """Calculate number of containers needed."""
    if total_pallets <= 0:
        return 0
    return ceil(total_pallets / PALLETS_PER_CONTAINER)


def get_production_month(boat_departure: date) -> str:
    """
    Get the production month name in Spanish.

    Production is typically for the month after boat departure.
    """
    # Add one month for production planning
    month = boat_departure.month + 1
    year = boat_departure.year

    if month > 12:
        month = 1
        year += 1

    return MONTHS_ES[month]


class ExportService:
    """Service for generating factory export files."""

    def generate_factory_order_excel(
        self,
        products: list[dict],
        boat_departure: date,
        order_date: Optional[date] = None,
    ) -> BytesIO:
        """
        Generate Excel file for factory order.

        Args:
            products: List of dicts with 'sku' and 'pallets' keys
            boat_departure: Boat departure date (used to calculate production month)
            order_date: Order date (defaults to today)

        Returns:
            BytesIO containing the Excel file
        """
        if order_date is None:
            order_date = date.today()

        production_month = get_production_month(boat_departure)

        logger.info(
            "generating_factory_order",
            product_count=len(products),
            boat_departure=str(boat_departure),
            production_month=production_month,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "PEDIDO TARRAGONA"

        # Styles
        bold_font = Font(bold=True)
        header_font = Font(bold=True, size=14)
        thin_border = Border(
            bottom=Side(style="thin", color="000000")
        )

        # Set column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 15

        # Row 1: Title
        ws["A1"] = "Pedido Tarragona Guatemala"
        ws["A1"].font = header_font

        # Row 3: Order date
        ws["A3"] = "Fecha de pedido:"
        ws["B3"] = order_date.strftime("%d/%m/%Y")

        # Row 5: Production month
        ws["A5"] = "Fabricacion para:"
        ws["B5"] = production_month
        ws["B5"].font = bold_font

        # Row 7: Column headers
        ws["A7"] = "Referencia"
        ws["B7"] = "Formato"
        ws["C7"] = "M2 solicitados"
        ws["A7"].font = bold_font
        ws["B7"].font = bold_font
        ws["C7"].font = bold_font
        ws["A7"].border = thin_border
        ws["B7"].border = thin_border
        ws["C7"].border = thin_border

        # Products (starting row 8)
        row = 8
        total_m2 = Decimal("0")
        total_pallets = 0

        for product in products:
            sku = product.get("sku", "")
            pallets = product.get("pallets", 0)

            if pallets <= 0:
                continue

            referencia = normalize_sku_for_factory(sku)
            formato = "51X51"  # Default format
            m2 = calculate_m2(pallets)

            ws[f"A{row}"] = referencia
            ws[f"B{row}"] = formato
            ws[f"C{row}"] = round(float(m2))
            ws[f"C{row}"].number_format = "#,##0"

            total_m2 += m2
            total_pallets += pallets
            row += 1

        # Empty row
        row += 1

        # Total row
        ws[f"A{row}"] = "TOTAL"
        ws[f"A{row}"].font = bold_font
        ws[f"C{row}"] = round(float(total_m2))
        ws[f"C{row}"].font = bold_font
        ws[f"C{row}"].number_format = "#,##0"
        ws[f"A{row}"].border = thin_border
        ws[f"B{row}"].border = thin_border
        ws[f"C{row}"].border = thin_border

        # Container count
        row += 2
        containers = calculate_containers(total_pallets)
        ws[f"A{row}"] = f"{containers} CONTENEDORES"
        ws[f"A{row}"].font = bold_font

        logger.info(
            "factory_order_generated",
            total_products=len(products),
            total_m2=float(total_m2),
            total_pallets=total_pallets,
            containers=containers,
        )

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    def generate_bl_allocation_excel(
        self,
        report: BLAllocationReport,
    ) -> BytesIO:
        """
        Generate Excel file for BL allocation report.

        Creates:
        - Summary sheet with risk distribution
        - One sheet per BL with product details

        Args:
            report: BLAllocationReport with allocation data

        Returns:
            BytesIO containing the Excel file
        """
        logger.info(
            "generating_bl_allocation_excel",
            num_bls=report.num_bls,
            total_products=sum(len(bl.products) for bl in report.allocations),
            total_critical=report.total_critical_products,
        )

        wb = Workbook()

        # Styles
        bold_font = Font(bold=True)
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=11)
        thin_border = Border(
            bottom=Side(style="thin", color="000000")
        )
        critical_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
        header_fill = PatternFill(start_color="E0E8FF", end_color="E0E8FF", fill_type="solid")
        success_fill = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")
        warning_fill = PatternFill(start_color="FFF0E0", end_color="FFF0E0", fill_type="solid")

        # ===== SUMMARY SHEET =====
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Column widths
        ws_summary.column_dimensions["A"].width = 35
        ws_summary.column_dimensions["B"].width = 20
        ws_summary.column_dimensions["C"].width = 15
        ws_summary.column_dimensions["D"].width = 15

        row = 1

        # Title
        ws_summary[f"A{row}"] = "BL ALLOCATION REPORT"
        ws_summary[f"A{row}"].font = title_font
        row += 2

        # Metadata
        ws_summary[f"A{row}"] = "Generated:"
        ws_summary[f"B{row}"] = report.generated_at.strftime("%Y-%m-%d %H:%M")
        row += 1
        ws_summary[f"A{row}"] = "Boat:"
        ws_summary[f"B{row}"] = report.boat_name
        row += 1
        ws_summary[f"A{row}"] = "Departure:"
        ws_summary[f"B{row}"] = report.boat_departure.strftime("%Y-%m-%d")
        row += 2

        # Totals
        ws_summary[f"A{row}"] = "TOTALS"
        ws_summary[f"A{row}"].font = header_font
        ws_summary[f"A{row}"].fill = header_fill
        ws_summary[f"B{row}"].fill = header_fill
        row += 1

        ws_summary[f"A{row}"] = "Total BLs:"
        ws_summary[f"B{row}"] = report.num_bls
        row += 1
        ws_summary[f"A{row}"] = "Total Containers:"
        ws_summary[f"B{row}"] = report.total_containers
        row += 1
        ws_summary[f"A{row}"] = "Total Pallets:"
        ws_summary[f"B{row}"] = report.total_pallets
        row += 1
        ws_summary[f"A{row}"] = "Total M2:"
        ws_summary[f"B{row}"] = round(float(report.total_m2))
        ws_summary[f"B{row}"].number_format = "#,##0"
        row += 1
        ws_summary[f"A{row}"] = "Total Weight (kg):"
        ws_summary[f"B{row}"] = round(float(report.total_weight_kg))
        ws_summary[f"B{row}"].number_format = "#,##0"
        row += 2

        # Risk Distribution
        ws_summary[f"A{row}"] = "RISK DISTRIBUTION"
        ws_summary[f"A{row}"].font = header_font
        if report.risk_distribution_even:
            ws_summary[f"A{row}"].fill = success_fill
            ws_summary[f"B{row}"].fill = success_fill
        else:
            ws_summary[f"A{row}"].fill = warning_fill
            ws_summary[f"B{row}"].fill = warning_fill
        row += 1

        ws_summary[f"A{row}"] = "Total Critical Products:"
        ws_summary[f"B{row}"] = report.total_critical_products
        row += 1

        # Distribution per BL
        for bl in report.allocations:
            if report.total_critical_products > 0:
                pct = (bl.critical_product_count / report.total_critical_products) * 100
            else:
                pct = 0
            ws_summary[f"A{row}"] = f"BL {bl.bl_number} Critical:"
            ws_summary[f"B{row}"] = f"{bl.critical_product_count} ({pct:.0f}%)"
            row += 1

        row += 1

        # Risk assessment
        ws_summary[f"A{row}"] = "Risk Status:"
        if report.risk_distribution_even:
            ws_summary[f"B{row}"] = "EVENLY DISTRIBUTED"
            ws_summary[f"B{row}"].font = Font(bold=True, color="006600")
        else:
            ws_summary[f"B{row}"] = f"UNEVEN ({report.max_critical_pct:.0f}% in one BL)"
            ws_summary[f"B{row}"].font = Font(bold=True, color="CC6600")
        row += 1

        ws_summary[f"A{row}"] = "Max delay if BL held:"
        ws_summary[f"B{row}"] = f"{report.max_critical_pct:.0f}% of critical products"
        row += 2

        # Warnings
        if report.warnings:
            ws_summary[f"A{row}"] = "WARNINGS"
            ws_summary[f"A{row}"].font = header_font
            ws_summary[f"A{row}"].fill = warning_fill
            row += 1
            for warning in report.warnings:
                ws_summary[f"A{row}"] = warning
                row += 1
            row += 1

        # BL Overview Table
        ws_summary[f"A{row}"] = "BL OVERVIEW"
        ws_summary[f"A{row}"].font = header_font
        ws_summary[f"A{row}"].fill = header_fill
        row += 1

        # Headers
        ws_summary[f"A{row}"] = "BL"
        ws_summary[f"B{row}"] = "Customers"
        ws_summary[f"C{row}"] = "Containers"
        ws_summary[f"D{row}"] = "Critical"
        for col in ["A", "B", "C", "D"]:
            ws_summary[f"{col}{row}"].font = bold_font
            ws_summary[f"{col}{row}"].border = thin_border
        row += 1

        for bl in report.allocations:
            customers = ", ".join(bl.primary_customers[:3])
            if len(bl.primary_customers) > 3:
                customers += f" +{len(bl.primary_customers) - 3}"
            ws_summary[f"A{row}"] = f"BL {bl.bl_number}"
            ws_summary[f"B{row}"] = customers or "General Stock"
            ws_summary[f"C{row}"] = bl.total_containers
            ws_summary[f"D{row}"] = bl.critical_product_count
            row += 1

        # ===== PER-BL SHEETS =====
        for bl in report.allocations:
            ws = wb.create_sheet(title=f"BL {bl.bl_number}")

            # Column widths
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 10
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 20
            ws.column_dimensions["E"].width = 10
            ws.column_dimensions["F"].width = 10

            row = 1

            # BL Header
            customers_str = ", ".join(bl.primary_customers[:5])
            if len(bl.primary_customers) > 5:
                customers_str += f" +{len(bl.primary_customers) - 5}"

            ws[f"A{row}"] = f"BL {bl.bl_number} — Serving: {customers_str or 'General Stock'}"
            ws[f"A{row}"].font = title_font
            row += 2

            # BL Stats
            ws[f"A{row}"] = f"Containers: {bl.total_containers}"
            ws[f"B{row}"] = f"Pallets: {bl.total_pallets}"
            ws[f"C{row}"] = f"Critical: {bl.critical_product_count}"
            row += 2

            # Product Headers
            headers = ["Referencia", "Pallets", "M2", "Customer", "Score", "Critical"]
            for i, header in enumerate(headers):
                col = chr(65 + i)  # A, B, C, ...
                ws[f"{col}{row}"] = header
                ws[f"{col}{row}"].font = bold_font
                ws[f"{col}{row}"].border = thin_border
                ws[f"{col}{row}"].fill = header_fill
            row += 1

            # Products
            for product in bl.products:
                ws[f"A{row}"] = normalize_sku_for_factory(product.sku)
                ws[f"B{row}"] = product.pallets
                ws[f"C{row}"] = round(float(product.m2))
                ws[f"C{row}"].number_format = "#,##0"
                ws[f"D{row}"] = product.primary_customer or "(General)"
                ws[f"E{row}"] = product.score

                if product.is_critical:
                    ws[f"F{row}"] = "YES"
                    ws[f"F{row}"].font = Font(bold=True, color="CC0000")
                    # Highlight entire row
                    for col in ["A", "B", "C", "D", "E", "F"]:
                        ws[f"{col}{row}"].fill = critical_fill
                else:
                    ws[f"F{row}"] = ""

                row += 1

            # Subtotal
            row += 1
            ws[f"A{row}"] = "SUBTOTAL"
            ws[f"A{row}"].font = bold_font
            ws[f"B{row}"] = bl.total_pallets
            ws[f"B{row}"].font = bold_font
            ws[f"C{row}"] = round(float(bl.total_m2))
            ws[f"C{row}"].font = bold_font
            ws[f"C{row}"].number_format = "#,##0"
            for col in ["A", "B", "C"]:
                ws[f"{col}{row}"].border = thin_border

        logger.info(
            "bl_allocation_excel_generated",
            num_bls=report.num_bls,
            total_containers=report.total_containers,
        )

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    def generate_order_report(
        self,
        response: OrderBuilderResponse,
        selected_warehouse_products: List[OrderBuilderProduct],
        selected_add_items: List[AddToProductionItem],
        selected_factory_items: List[FactoryRequestItem],
    ) -> BytesIO:
        """
        Generate comprehensive Order Builder report explaining WHY recommendations were made.

        Creates Excel with 7 sheets:
        1. EXECUTIVE SUMMARY - High-level order overview with inventory impact
        2. SHIP NOW - Products to ship from SIESA warehouse
        3. CALCULATION BREAKDOWN - Full per-product calculations
        4. ADD TO PRODUCTION - Items to add to scheduled production (no deadline)
        5. FACTORY REQUEST - New production requests (1 container minimum)
        6. SELECTION CRITERIA - Why products were selected
        7. NOT INCLUDED - Products NOT shipped and reasons

        Args:
            response: Full OrderBuilderResponse from the API
            selected_warehouse_products: Products selected for warehouse order
            selected_add_items: Items selected for add to production
            selected_factory_items: Items selected for factory request

        Returns:
            BytesIO containing the Excel file
        """
        logger.info(
            "generating_order_report",
            warehouse_count=len(selected_warehouse_products),
            add_count=len(selected_add_items),
            factory_count=len(selected_factory_items),
        )

        wb = Workbook()

        # Styles
        bold_font = Font(bold=True)
        title_font = Font(bold=True, size=16)
        section_font = Font(bold=True, size=12)
        subtitle_font = Font(bold=True, size=11)
        thin_border = Border(bottom=Side(style="thin", color="000000"))
        header_fill = PatternFill(start_color="E0E8FF", end_color="E0E8FF", fill_type="solid")
        critical_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
        urgent_fill = PatternFill(start_color="FFF0E0", end_color="FFF0E0", fill_type="solid")
        ok_fill = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")
        gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

        # Calculate totals upfront
        ship_now_pallets = sum(p.selected_pallets for p in selected_warehouse_products)
        ship_now_m2 = Decimal(str(ship_now_pallets)) * M2_PER_PALLET_FACTORY
        ship_now_containers = ceil(ship_now_pallets / PALLETS_PER_CONTAINER)

        add_pallets = sum(item.suggested_additional_pallets for item in selected_add_items)
        add_m2 = sum(item.suggested_additional_m2 for item in selected_add_items)

        factory_pallets = sum(item.request_pallets for item in selected_factory_items)
        factory_m2 = sum(item.request_m2 for item in selected_factory_items)

        # Get all products for NOT INCLUDED sheet
        all_products = (
            response.high_priority +
            response.consider +
            response.well_covered +
            response.your_call
        )
        selected_ids = {p.product_id for p in selected_warehouse_products}

        # Helper for target boat formatting: "Feb 19 — BOAT NAME"
        def format_target_boat(boat_name: str, boat_departure: date = None) -> str:
            if boat_departure:
                date_str = boat_departure.strftime("%b %d")  # "Feb 19"
                if boat_name:
                    return f"{date_str} — {boat_name}"
                return date_str
            elif boat_name:
                return boat_name
            return "TBD"

        # ===== SHEET 1: EXECUTIVE SUMMARY =====
        ws_summary = wb.active
        ws_summary.title = "EXECUTIVE SUMMARY"

        ws_summary.column_dimensions["A"].width = 40
        ws_summary.column_dimensions["B"].width = 30
        ws_summary.column_dimensions["C"].width = 25

        row = 1
        ws_summary[f"A{row}"] = "ORDER BUILDER REPORT"
        ws_summary[f"A{row}"].font = title_font
        row += 1
        ws_summary[f"A{row}"] = f"Generated: {date.today().strftime('%Y-%m-%d')}"
        row += 2

        # === BOAT INFORMATION ===
        ws_summary[f"A{row}"] = "TARGET BOAT"
        ws_summary[f"A{row}"].font = section_font
        ws_summary[f"A{row}"].fill = header_fill
        ws_summary[f"B{row}"].fill = header_fill
        row += 1

        # Format boat display
        boat_display = format_target_boat(response.boat.name, response.boat.departure_date)
        ws_summary[f"A{row}"] = "Boat:"
        ws_summary[f"B{row}"] = boat_display
        ws_summary[f"B{row}"].font = bold_font
        row += 1
        ws_summary[f"A{row}"] = "Departure Date:"
        ws_summary[f"B{row}"] = response.boat.departure_date.strftime("%Y-%m-%d")
        row += 1
        ws_summary[f"A{row}"] = "Days Until Departure:"
        ws_summary[f"B{row}"] = response.boat.days_until_departure
        row += 1

        # Order Deadline (30 days before departure)
        ws_summary[f"A{row}"] = "Order Deadline:"
        ws_summary[f"B{row}"] = response.boat.order_deadline.strftime("%Y-%m-%d")
        row += 1
        ws_summary[f"A{row}"] = "Days Until Order Deadline:"
        ws_summary[f"B{row}"] = response.boat.days_until_order_deadline
        if response.boat.days_until_order_deadline <= 7:
            ws_summary[f"B{row}"].fill = critical_fill
        elif response.boat.days_until_order_deadline <= 14:
            ws_summary[f"B{row}"].fill = urgent_fill
        row += 1

        if response.boat.past_order_deadline:
            ws_summary[f"A{row}"] = "⚠️ PAST ORDER DEADLINE"
            ws_summary[f"A{row}"].font = Font(bold=True, color="CC0000")
            row += 1
        row += 1

        # === ORDER SUMMARY ===
        ws_summary[f"A{row}"] = "ORDER SUMMARY"
        ws_summary[f"A{row}"].font = section_font
        ws_summary[f"A{row}"].fill = header_fill
        ws_summary[f"B{row}"].fill = header_fill
        row += 1

        ws_summary[f"A{row}"] = "Ship Now (SIESA → Warehouse):"
        ws_summary[f"B{row}"] = f"{ship_now_m2:,.0f} m² ({ship_now_pallets} pallets, {ship_now_containers} containers)"
        row += 1
        ws_summary[f"A{row}"] = "Add to Production:"
        ws_summary[f"B{row}"] = f"{add_m2:,.0f} m² ({add_pallets} pallets)"
        row += 1
        ws_summary[f"A{row}"] = "New Factory Request:"
        ws_summary[f"B{row}"] = f"{factory_m2:,.0f} m² ({factory_pallets} pallets)"
        row += 1

        total_order_m2 = ship_now_m2 + add_m2 + factory_m2
        total_order_pallets = ship_now_pallets + add_pallets + factory_pallets
        ws_summary[f"A{row}"] = "TOTAL ORDER:"
        ws_summary[f"A{row}"].font = bold_font
        ws_summary[f"B{row}"] = f"{total_order_m2:,.0f} m² ({total_order_pallets} pallets)"
        ws_summary[f"B{row}"].font = bold_font
        row += 2

        # === BL ALLOCATION ===
        ws_summary[f"A{row}"] = "BL ALLOCATION"
        ws_summary[f"A{row}"].font = section_font
        ws_summary[f"A{row}"].fill = header_fill
        ws_summary[f"B{row}"].fill = header_fill
        row += 1

        ws_summary[f"A{row}"] = "Selected BLs:"
        ws_summary[f"B{row}"] = response.num_bls
        row += 1
        ws_summary[f"A{row}"] = "Recommended BLs (based on need):"
        ws_summary[f"B{row}"] = response.recommended_bls
        row += 1
        ws_summary[f"A{row}"] = "Available BLs (based on SIESA stock):"
        ws_summary[f"B{row}"] = response.available_bls
        row += 1
        ws_summary[f"A{row}"] = "BL Capacity:"
        ws_summary[f"B{row}"] = f"{response.num_bls * 70} pallets ({response.num_bls * 5} containers)"
        row += 1

        if response.recommended_bls_reason:
            ws_summary[f"A{row}"] = "Recommendation:"
            ws_summary[f"B{row}"] = response.recommended_bls_reason
            row += 1
        row += 1

        # === URGENCY BREAKDOWN ===
        ws_summary[f"A{row}"] = "URGENCY BREAKDOWN"
        ws_summary[f"A{row}"].font = section_font
        ws_summary[f"A{row}"].fill = header_fill
        ws_summary[f"B{row}"].fill = header_fill
        row += 1

        critical_count = len([p for p in selected_warehouse_products if p.score and p.score.total >= 85])
        urgent_count = len([p for p in selected_warehouse_products if p.score and 70 <= p.score.total < 85])
        other_count = len(selected_warehouse_products) - critical_count - urgent_count

        ws_summary[f"A{row}"] = "Critical (Score ≥ 85):"
        ws_summary[f"B{row}"] = f"{critical_count} products"
        ws_summary[f"B{row}"].fill = critical_fill
        row += 1
        ws_summary[f"A{row}"] = "Urgent (Score 70-84):"
        ws_summary[f"B{row}"] = f"{urgent_count} products"
        ws_summary[f"B{row}"].fill = urgent_fill
        row += 1
        ws_summary[f"A{row}"] = "Standard:"
        ws_summary[f"B{row}"] = f"{other_count} products"
        row += 2

        # === INVENTORY IMPACT ===
        ws_summary[f"A{row}"] = "INVENTORY IMPACT"
        ws_summary[f"A{row}"].font = section_font
        ws_summary[f"A{row}"].fill = header_fill
        ws_summary[f"B{row}"].fill = header_fill
        row += 1

        ws_summary[f"A{row}"] = "Current Warehouse Utilization:"
        current_util = response.summary.warehouse_current_pallets
        ws_summary[f"B{row}"] = f"{current_util} / {response.summary.warehouse_capacity} pallets"
        row += 1

        ws_summary[f"A{row}"] = "After Ship Now Delivery:"
        ws_summary[f"B{row}"] = f"{response.summary.warehouse_after_delivery} pallets ({response.summary.warehouse_utilization_after:.0f}%)"
        if response.summary.warehouse_utilization_after > 95:
            ws_summary[f"B{row}"].fill = urgent_fill
        row += 2

        # === STRATEGY SUMMARY ===
        if response.summary_reasoning and response.summary_reasoning.reasoning:
            reasoning = response.summary_reasoning.reasoning
            ws_summary[f"A{row}"] = "STRATEGY SUMMARY"
            ws_summary[f"A{row}"].font = section_font
            ws_summary[f"A{row}"].fill = header_fill
            ws_summary[f"B{row}"].fill = header_fill
            row += 1

            ws_summary[f"A{row}"] = "Strategy:"
            ws_summary[f"B{row}"] = reasoning.strategy_sentence
            row += 1
            ws_summary[f"A{row}"] = "Risk:"
            ws_summary[f"B{row}"] = reasoning.risk_sentence
            row += 1
            ws_summary[f"A{row}"] = "Constraint:"
            ws_summary[f"B{row}"] = reasoning.constraint_sentence
            row += 1
            if reasoning.customer_sentence:
                ws_summary[f"A{row}"] = "Customer Signal:"
                ws_summary[f"B{row}"] = reasoning.customer_sentence
                row += 1

        # ===== SHEET 2: SHIP NOW =====
        ws_ship = wb.create_sheet(title="SHIP NOW")
        ws_ship.column_dimensions["A"].width = 30
        ws_ship.column_dimensions["B"].width = 12
        ws_ship.column_dimensions["C"].width = 12
        ws_ship.column_dimensions["D"].width = 15
        ws_ship.column_dimensions["E"].width = 12
        ws_ship.column_dimensions["F"].width = 15

        row = 1
        ws_ship[f"A{row}"] = "SHIP NOW - WAREHOUSE ORDER"
        ws_ship[f"A{row}"].font = title_font
        row += 1
        ws_ship[f"A{row}"] = f"Target: {format_target_boat(response.boat.name, response.boat.departure_date)}"
        row += 2

        # Headers
        headers = ["SKU", "Pallets", "M²", "SIESA Stock", "Score", "Urgency"]
        for i, header in enumerate(headers):
            col = chr(65 + i)
            ws_ship[f"{col}{row}"] = header
            ws_ship[f"{col}{row}"].font = bold_font
            ws_ship[f"{col}{row}"].border = thin_border
            ws_ship[f"{col}{row}"].fill = header_fill
        row += 1

        # Products
        for p in selected_warehouse_products:
            ws_ship[f"A{row}"] = p.sku
            ws_ship[f"B{row}"] = p.selected_pallets
            ws_ship[f"C{row}"] = round(float(Decimal(p.selected_pallets) * M2_PER_PALLET_FACTORY))
            ws_ship[f"C{row}"].number_format = "#,##0"
            ws_ship[f"D{row}"] = round(float(p.factory_available_m2))
            ws_ship[f"D{row}"].number_format = "#,##0"
            ws_ship[f"E{row}"] = p.score.total if p.score else 0
            ws_ship[f"F{row}"] = p.urgency.upper()

            if p.urgency == "critical":
                for col in ["A", "B", "C", "D", "E", "F"]:
                    ws_ship[f"{col}{row}"].fill = critical_fill
            elif p.urgency == "urgent":
                for col in ["A", "B", "C", "D", "E", "F"]:
                    ws_ship[f"{col}{row}"].fill = urgent_fill

            row += 1

        # Totals
        row += 1
        ws_ship[f"A{row}"] = "TOTAL"
        ws_ship[f"A{row}"].font = bold_font
        ws_ship[f"B{row}"] = ship_now_pallets
        ws_ship[f"B{row}"].font = bold_font
        ws_ship[f"C{row}"] = round(float(ship_now_m2))
        ws_ship[f"C{row}"].font = bold_font
        ws_ship[f"C{row}"].number_format = "#,##0"

        # ===== SHEET 3: CALCULATION BREAKDOWN =====
        ws_calc = wb.create_sheet(title="CALCULATION BREAKDOWN")
        ws_calc.column_dimensions["A"].width = 25
        ws_calc.column_dimensions["B"].width = 12
        ws_calc.column_dimensions["C"].width = 12
        ws_calc.column_dimensions["D"].width = 12
        ws_calc.column_dimensions["E"].width = 12
        ws_calc.column_dimensions["F"].width = 12
        ws_calc.column_dimensions["G"].width = 12
        ws_calc.column_dimensions["H"].width = 12
        ws_calc.column_dimensions["I"].width = 12
        ws_calc.column_dimensions["J"].width = 12

        row = 1
        ws_calc[f"A{row}"] = "CALCULATION BREAKDOWN"
        ws_calc[f"A{row}"].font = title_font
        row += 1
        ws_calc[f"A{row}"] = "Full calculation details for each product"
        row += 2

        # Headers
        headers = [
            "SKU", "Velocity\n(m²/day)", "Lead Time\n(days)", "Coverage\n(days)",
            "Base Need\n(m²)", "Trend Adj\n(%)", "Current\nStock (m²)",
            "In Transit\n(m²)", "Final Gap\n(m²)", "Pallets"
        ]
        for i, header in enumerate(headers):
            col = chr(65 + i)
            ws_calc[f"{col}{row}"] = header
            ws_calc[f"{col}{row}"].font = bold_font
            ws_calc[f"{col}{row}"].border = thin_border
            ws_calc[f"{col}{row}"].fill = header_fill
            ws_calc[f"{col}{row}"].alignment = Alignment(wrap_text=True)
        row += 1

        # Product calculations
        for p in selected_warehouse_products:
            ws_calc[f"A{row}"] = p.sku

            if p.calculation_breakdown:
                cb = p.calculation_breakdown
                ws_calc[f"B{row}"] = round(float(cb.daily_velocity_m2), 1)
                ws_calc[f"C{row}"] = cb.lead_time_days
                ws_calc[f"D{row}"] = cb.lead_time_days + cb.ordering_cycle_days
                ws_calc[f"E{row}"] = round(float(cb.base_quantity_m2))
                ws_calc[f"E{row}"].number_format = "#,##0"
                ws_calc[f"F{row}"] = f"{cb.trend_adjustment_pct:+.0f}%" if cb.trend_adjustment_pct else "0%"
                ws_calc[f"G{row}"] = round(float(cb.minus_current_stock_m2))
                ws_calc[f"G{row}"].number_format = "#,##0"
                ws_calc[f"H{row}"] = round(float(cb.minus_incoming_m2))
                ws_calc[f"H{row}"].number_format = "#,##0"
                ws_calc[f"I{row}"] = round(float(cb.final_suggestion_m2))
                ws_calc[f"I{row}"].number_format = "#,##0"
                ws_calc[f"J{row}"] = cb.final_suggestion_pallets
            else:
                # Fallback for products without breakdown
                ws_calc[f"B{row}"] = round(float(p.daily_velocity_m2), 1) if p.daily_velocity_m2 else 0
                ws_calc[f"C{row}"] = p.days_to_cover
                ws_calc[f"D{row}"] = p.days_to_cover + 30  # Assume 30 day cycle
                ws_calc[f"E{row}"] = round(float(p.total_demand_m2))
                ws_calc[f"F{row}"] = "N/A"
                ws_calc[f"G{row}"] = round(float(p.current_stock_m2))
                ws_calc[f"H{row}"] = round(float(p.in_transit_m2))
                ws_calc[f"I{row}"] = round(float(p.coverage_gap_m2))
                ws_calc[f"J{row}"] = p.selected_pallets

            if p.urgency == "critical":
                for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
                    ws_calc[f"{col}{row}"].fill = critical_fill
            elif p.urgency == "urgent":
                for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
                    ws_calc[f"{col}{row}"].fill = urgent_fill

            row += 1

        # Formula explanation
        row += 2
        ws_calc[f"A{row}"] = "CALCULATION FORMULA"
        ws_calc[f"A{row}"].font = section_font
        row += 1
        ws_calc[f"A{row}"] = "Base Need = Velocity × (Lead Time + Coverage Days)"
        row += 1
        ws_calc[f"A{row}"] = "Final Gap = Base Need + Trend Adjustment - Current Stock - In Transit"
        row += 1
        ws_calc[f"A{row}"] = "Pallets = ceil(Final Gap ÷ 134.4 m²/pallet)"

        # ===== SHEET 4: ADD TO PRODUCTION =====
        ws_add = wb.create_sheet(title="ADD TO PRODUCTION")
        ws_add.column_dimensions["A"].width = 30
        ws_add.column_dimensions["B"].width = 15
        ws_add.column_dimensions["C"].width = 15
        ws_add.column_dimensions["D"].width = 15
        ws_add.column_dimensions["E"].width = 25
        ws_add.column_dimensions["F"].width = 12

        row = 1
        ws_add[f"A{row}"] = "ADD TO PRODUCTION"
        ws_add[f"A{row}"].font = title_font
        row += 1
        ws_add[f"A{row}"] = "Items already in production schedule — can add more before production starts"
        row += 1
        ws_add[f"A{row}"] = "No minimum quantity required for additions"
        ws_add[f"A{row}"].font = Font(italic=True)
        row += 2

        # Headers
        headers = ["SKU", "Current Sched.", "Add This", "New Total", "Target Boat", "Score"]
        for i, header in enumerate(headers):
            col = chr(65 + i)
            ws_add[f"{col}{row}"] = header
            ws_add[f"{col}{row}"].font = bold_font
            ws_add[f"{col}{row}"].border = thin_border
            ws_add[f"{col}{row}"].fill = header_fill
        row += 1

        # Items
        for item in selected_add_items:
            ws_add[f"A{row}"] = item.sku
            ws_add[f"B{row}"] = f"{item.current_requested_m2:,.0f} m²"
            ws_add[f"C{row}"] = f"{item.suggested_additional_m2:,.0f} m²"
            ws_add[f"D{row}"] = f"{item.suggested_total_m2:,.0f} m²"

            # Format target boat with date
            if item.target_boat_departure:
                ws_add[f"E{row}"] = format_target_boat(item.target_boat, item.target_boat_departure)
            elif item.target_boat:
                ws_add[f"E{row}"] = item.target_boat
            else:
                # Show estimated ready date if no boat matched
                if item.estimated_ready_date:
                    ws_add[f"E{row}"] = f"TBD — ready ~{item.estimated_ready_date.strftime('%b %d')}"
                else:
                    ws_add[f"E{row}"] = "TBD"

            ws_add[f"F{row}"] = item.score

            if item.is_critical:
                for col in ["A", "B", "C", "D", "E", "F"]:
                    ws_add[f"{col}{row}"].fill = critical_fill

            row += 1

        # Totals
        row += 1
        ws_add[f"A{row}"] = "TOTAL TO ADD"
        ws_add[f"A{row}"].font = bold_font
        ws_add[f"C{row}"] = f"{add_m2:,.0f} m²"
        ws_add[f"C{row}"].font = bold_font

        # ===== SHEET 5: FACTORY REQUEST =====
        ws_factory = wb.create_sheet(title="FACTORY REQUEST")
        ws_factory.column_dimensions["A"].width = 30
        ws_factory.column_dimensions["B"].width = 15
        ws_factory.column_dimensions["C"].width = 12
        ws_factory.column_dimensions["D"].width = 15
        ws_factory.column_dimensions["E"].width = 15
        ws_factory.column_dimensions["F"].width = 12
        ws_factory.column_dimensions["G"].width = 15

        row = 1
        ws_factory[f"A{row}"] = "NEW FACTORY REQUEST"
        ws_factory[f"A{row}"].font = title_font
        row += 1
        ws_factory[f"A{row}"] = "Items NOT in current production schedule"
        row += 1
        ws_factory[f"A{row}"] = "⚠️ Minimum order: 1 container (14 pallets / 1,881.6 m²)"
        ws_factory[f"A{row}"].font = Font(bold=True, color="CC6600")
        row += 2

        # Quota tracking
        if response.factory_request_summary:
            frs = response.factory_request_summary
            ws_factory[f"A{row}"] = "Monthly Quota:"
            ws_factory[f"B{row}"] = f"{frs.limit_m2:,.0f} m²"
            row += 1
            ws_factory[f"A{row}"] = "Already Requested:"
            ws_factory[f"B{row}"] = f"{frs.limit_m2 - frs.remaining_m2:,.0f} m² ({frs.utilization_pct:.0f}%)"
            row += 1
            ws_factory[f"A{row}"] = "Remaining:"
            ws_factory[f"B{row}"] = f"{frs.remaining_m2:,.0f} m²"
            if frs.remaining_m2 < factory_m2:
                ws_factory[f"B{row}"].fill = urgent_fill
            row += 1
            if frs.estimated_ready:
                ws_factory[f"A{row}"] = "Estimated Ready:"
                ws_factory[f"B{row}"] = frs.estimated_ready
                row += 1
            row += 1

        # Headers
        headers = ["SKU", "Request M²", "Pallets", "SIESA Stock", "In Transit", "Urgency", "Note"]
        for i, header in enumerate(headers):
            col = chr(65 + i)
            ws_factory[f"{col}{row}"] = header
            ws_factory[f"{col}{row}"].font = bold_font
            ws_factory[f"{col}{row}"].border = thin_border
            ws_factory[f"{col}{row}"].fill = header_fill
        row += 1

        # Items with minimum container rule applied
        MIN_PALLETS = 14
        MIN_M2 = Decimal("1881.6")

        for item in selected_factory_items:
            ws_factory[f"A{row}"] = item.sku

            # Apply minimum 1 container rule for NEW requests
            actual_request_pallets = max(item.request_pallets, MIN_PALLETS)
            actual_request_m2 = max(item.request_m2, MIN_M2)

            ws_factory[f"B{row}"] = round(float(actual_request_m2))
            ws_factory[f"B{row}"].number_format = "#,##0"
            ws_factory[f"C{row}"] = actual_request_pallets
            ws_factory[f"D{row}"] = round(float(item.factory_available_m2))
            ws_factory[f"D{row}"].number_format = "#,##0"
            ws_factory[f"E{row}"] = round(float(item.in_transit_m2))
            ws_factory[f"E{row}"].number_format = "#,##0"
            ws_factory[f"F{row}"] = item.urgency.upper()

            # Note if rounded up to minimum
            if item.request_pallets < MIN_PALLETS:
                ws_factory[f"G{row}"] = f"Rounded up from {item.gap_pallets} pallets"
                ws_factory[f"G{row}"].font = Font(italic=True)
            else:
                ws_factory[f"G{row}"] = ""

            if item.urgency == "critical":
                for col in ["A", "B", "C", "D", "E", "F"]:
                    ws_factory[f"{col}{row}"].fill = critical_fill
            elif item.urgency == "urgent":
                for col in ["A", "B", "C", "D", "E", "F"]:
                    ws_factory[f"{col}{row}"].fill = urgent_fill

            row += 1

        # Totals (with minimum applied)
        row += 1
        total_factory_pallets_adjusted = sum(max(item.request_pallets, MIN_PALLETS) for item in selected_factory_items)
        total_factory_m2_adjusted = Decimal(total_factory_pallets_adjusted) * M2_PER_PALLET_FACTORY

        ws_factory[f"A{row}"] = "TOTAL REQUEST"
        ws_factory[f"A{row}"].font = bold_font
        ws_factory[f"B{row}"] = round(float(total_factory_m2_adjusted))
        ws_factory[f"B{row}"].font = bold_font
        ws_factory[f"B{row}"].number_format = "#,##0"
        ws_factory[f"C{row}"] = total_factory_pallets_adjusted
        ws_factory[f"C{row}"].font = bold_font

        # ===== SHEET 6: SELECTION CRITERIA =====
        ws_criteria = wb.create_sheet(title="SELECTION CRITERIA")
        ws_criteria.column_dimensions["A"].width = 25
        ws_criteria.column_dimensions["B"].width = 15
        ws_criteria.column_dimensions["C"].width = 45
        ws_criteria.column_dimensions["D"].width = 18
        ws_criteria.column_dimensions["E"].width = 15

        row = 1
        ws_criteria[f"A{row}"] = "SELECTION CRITERIA"
        ws_criteria[f"A{row}"].font = title_font
        row += 1
        ws_criteria[f"A{row}"] = "Why each product was selected for this order"
        row += 2

        # Headers
        headers = ["SKU", "Priority Tier", "Why Selected", "Dominant Factor", "Score"]
        for i, header in enumerate(headers):
            col = chr(65 + i)
            ws_criteria[f"{col}{row}"] = header
            ws_criteria[f"{col}{row}"].font = bold_font
            ws_criteria[f"{col}{row}"].border = thin_border
            ws_criteria[f"{col}{row}"].fill = header_fill
        row += 1

        # Products with reasoning
        for p in selected_warehouse_products:
            ws_criteria[f"A{row}"] = p.sku
            ws_criteria[f"B{row}"] = p.priority

            if p.reasoning_display:
                ws_criteria[f"C{row}"] = p.reasoning_display.why_product_sentence
                ws_criteria[f"D{row}"] = p.reasoning_display.dominant_factor.upper()
            else:
                # Fallback reasoning
                days = p.days_of_stock or 0
                if days <= 0:
                    ws_criteria[f"C{row}"] = "OUT OF STOCK — immediate restock needed"
                    ws_criteria[f"D{row}"] = "STOCKOUT"
                elif days <= 14:
                    ws_criteria[f"C{row}"] = f"LOW STOCK — only {days} days remaining"
                    ws_criteria[f"D{row}"] = "STOCKOUT"
                else:
                    ws_criteria[f"C{row}"] = f"Coverage gap — {days} days of stock"
                    ws_criteria[f"D{row}"] = "COVERAGE"

            ws_criteria[f"E{row}"] = p.score.total if p.score else 0

            if p.urgency == "critical":
                for col in ["A", "B", "C", "D", "E"]:
                    ws_criteria[f"{col}{row}"].fill = critical_fill
            elif p.urgency == "urgent":
                for col in ["A", "B", "C", "D", "E"]:
                    ws_criteria[f"{col}{row}"].fill = urgent_fill

            row += 1

        # Scoring explanation
        row += 2
        ws_criteria[f"A{row}"] = "SCORING BREAKDOWN"
        ws_criteria[f"A{row}"].font = section_font
        row += 1
        ws_criteria[f"A{row}"] = "Stockout Risk (0-40 pts):"
        ws_criteria[f"B{row}"] = "Days of stock remaining"
        row += 1
        ws_criteria[f"A{row}"] = "Customer Demand (0-30 pts):"
        ws_criteria[f"B{row}"] = "Customers expected to order"
        row += 1
        ws_criteria[f"A{row}"] = "Growth Trend (0-20 pts):"
        ws_criteria[f"B{row}"] = "Velocity change %"
        row += 1
        ws_criteria[f"A{row}"] = "Revenue Impact (0-10 pts):"
        ws_criteria[f"B{row}"] = "Sales velocity"

        # ===== SHEET 7: NOT INCLUDED =====
        ws_not = wb.create_sheet(title="NOT INCLUDED")
        ws_not.column_dimensions["A"].width = 25
        ws_not.column_dimensions["B"].width = 15
        ws_not.column_dimensions["C"].width = 15
        ws_not.column_dimensions["D"].width = 15
        ws_not.column_dimensions["E"].width = 40

        row = 1
        ws_not[f"A{row}"] = "NOT INCLUDED IN THIS ORDER"
        ws_not[f"A{row}"].font = title_font
        row += 1
        ws_not[f"A{row}"] = "Products analyzed but not selected for shipping"
        row += 2

        # Headers
        headers = ["SKU", "Priority Tier", "Days of Stock", "Score", "Reason Not Included"]
        for i, header in enumerate(headers):
            col = chr(65 + i)
            ws_not[f"{col}{row}"] = header
            ws_not[f"{col}{row}"].font = bold_font
            ws_not[f"{col}{row}"].border = thin_border
            ws_not[f"{col}{row}"].fill = header_fill
        row += 1

        # Products NOT selected
        not_selected = [p for p in all_products if p.product_id not in selected_ids]

        for p in not_selected:
            ws_not[f"A{row}"] = p.sku
            ws_not[f"B{row}"] = p.priority
            ws_not[f"C{row}"] = p.days_of_stock if p.days_of_stock is not None else "N/A"
            ws_not[f"D{row}"] = p.score.total if p.score else 0

            # Determine reason not included
            reason = ""
            if p.priority == "WELL_COVERED":
                reason = f"Well stocked — {p.days_of_stock or 'N/A'} days of coverage"
            elif p.priority == "YOUR_CALL":
                if p.reasoning and p.reasoning.exclusion_reason:
                    exclusion = p.reasoning.exclusion_reason
                    if exclusion == "OVERSTOCKED":
                        reason = "Overstocked — declining demand + high inventory"
                    elif exclusion == "NO_SALES":
                        reason = "No recent sales in 90+ days"
                    elif exclusion == "DECLINING":
                        reason = "Significantly declining demand"
                    elif exclusion == "NO_DATA":
                        reason = "Insufficient sales data"
                    else:
                        reason = exclusion
                else:
                    reason = "Low priority — optional restock"
            elif p.in_transit_m2 and p.in_transit_m2 > 0:
                reason = f"In-transit covers need — {p.in_transit_m2:,.0f} m² on the way"
            elif p.factory_available_m2 and p.factory_available_m2 <= 0:
                reason = "No SIESA stock available"
            elif p.suggested_pallets <= 0:
                reason = "No coverage gap — current stock sufficient"
            else:
                reason = "Not selected — capacity constraint or lower priority"

            ws_not[f"E{row}"] = reason
            ws_not[f"E{row}"].fill = gray_fill

            row += 1

        if not not_selected:
            ws_not[f"A{row}"] = "All analyzed products were included in this order."
            ws_not[f"A{row}"].font = Font(italic=True)

        logger.info(
            "order_report_generated",
            warehouse_count=len(selected_warehouse_products),
            add_count=len(selected_add_items),
            factory_count=len(selected_factory_items),
            not_included_count=len(not_selected),
        )

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output


# Singleton instance
_export_service: Optional[ExportService] = None


def get_export_service() -> ExportService:
    """Get or create ExportService instance."""
    global _export_service
    if _export_service is None:
        _export_service = ExportService()
    return _export_service
