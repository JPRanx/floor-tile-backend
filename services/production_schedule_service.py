"""
Production Schedule service for CRUD operations.

See STANDARDS_LOGGING.md for logging patterns.
See STANDARDS_ERRORS.md for error handling patterns.
"""

from typing import Optional
from datetime import date, timedelta
from decimal import Decimal
from collections import defaultdict
import structlog

from config import get_supabase_client, settings
from models.production_schedule import (
    ParsedProductionSchedule,
    ProductionScheduleResponse,
    UpcomingProductionItem,
    # Order Builder integration
    FactoryStatus,
    ProductFactoryStatus,
    UnmappedProduct,
    MatchSuggestion,
    UploadResult,
    # Excel-based parsing
    ProductionStatus,
    ProductionScheduleCreate,
    ProductionScheduleDBResponse,
    ProductionSummary,
    ProductionCapacity,
    ProductionImportResult,
    CanAddMoreAlert,
)
from services.product_service import get_product_service
from exceptions import DatabaseError

# Excel parsing
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Optional: fuzzy matching
try:
    from rapidfuzz import fuzz
    FUZZY_AVAILABLE = True
except ImportError:
    FUZZY_AVAILABLE = False

logger = structlog.get_logger(__name__)

# Accent normalization map for Spanish characters
ACCENT_MAP = {
    'Á': 'A', 'á': 'a',
    'É': 'E', 'é': 'e',
    'Í': 'I', 'í': 'i',
    'Ó': 'O', 'ó': 'o',
    'Ú': 'U', 'ú': 'u',
    'Ñ': 'N', 'ñ': 'n',
    'Ü': 'U', 'ü': 'u',
}


def normalize_accents(text: str) -> str:
    """
    Remove Spanish accents for matching purposes.

    Examples:
        NOGAL CAFÉ BTE → NOGAL CAFE BTE
        CARACOLÍ → CARACOLI
        ALMENDRO MARRÓN → ALMENDRO MARRON
    """
    if not text:
        return text
    for accented, normalized in ACCENT_MAP.items():
        text = text.replace(accented, normalized)
    return text


def parse_plant_number(plant_value) -> int:
    """
    Parse plant value from database to integer.

    Handles:
        - Integer: 1 -> 1
        - String integer: "1" -> 1
        - Prefixed string: "plant_1" -> 1

    Returns 1 as default if parsing fails.
    """
    if plant_value is None:
        return 1
    if isinstance(plant_value, int):
        return plant_value
    if isinstance(plant_value, str):
        # Handle "plant_1", "plant_2" format
        if plant_value.startswith("plant_"):
            try:
                return int(plant_value.replace("plant_", ""))
            except ValueError:
                return 1
        # Handle plain string number
        try:
            return int(plant_value)
        except ValueError:
            return 1
    return 1


class ProductionScheduleService:
    """
    Production schedule business logic.

    Handles saving parsed schedules and querying upcoming production.
    """

    def __init__(self):
        self.db = get_supabase_client()
        self.table = "production_schedule"
        self.product_service = get_product_service()

    # ===================
    # UPSERT OPERATIONS
    # ===================

    def save_parsed_schedule(
        self,
        parsed_data: ParsedProductionSchedule,
        filename: Optional[str] = None
    ) -> tuple[int, int, list[str]]:
        """
        Save parsed production schedule to database.

        Uses upsert to handle re-uploading same schedule with updates.
        Matches factory codes to products where possible.

        Args:
            parsed_data: Parsed schedule from Claude Vision
            filename: Original PDF filename

        Returns:
            Tuple of (items_saved, products_matched, unmatched_factory_codes)
        """
        logger.info(
            "saving_production_schedule",
            schedule_date=str(parsed_data.schedule_date),
            line_items_count=len(parsed_data.line_items),
            filename=filename
        )

        items_saved = 0
        products_matched = 0
        unmatched_codes = set()

        # Get all unique factory codes and batch lookup products
        factory_codes = list(set(item.factory_code for item in parsed_data.line_items))
        products_by_code = {}

        if factory_codes:
            products = self.product_service.get_by_factory_codes(factory_codes)
            products_by_code = {p.factory_code: p for p in products if p.factory_code}

        for item in parsed_data.line_items:
            try:
                # Match product by factory code
                product = products_by_code.get(item.factory_code)
                product_id = product.id if product else None

                if product:
                    products_matched += 1
                else:
                    unmatched_codes.add(item.factory_code)

                # Build upsert data
                upsert_data = {
                    "schedule_date": parsed_data.schedule_date.isoformat(),
                    "schedule_version": parsed_data.schedule_version,
                    "source_filename": filename,
                    "production_date": item.production_date.isoformat(),
                    "factory_code": item.factory_code,
                    "product_name": item.product_name,
                    "product_id": product_id,
                    "plant": item.plant,
                    "format": item.format,
                    "design": item.design,
                    "finish": item.finish,
                    "shifts": float(item.shifts) if item.shifts else None,
                    "quality_target_pct": float(item.quality_target_pct) if item.quality_target_pct else None,
                    "quality_actual_pct": float(item.quality_actual_pct) if item.quality_actual_pct else None,
                    "m2_total_net": float(item.m2_total_net) if item.m2_total_net else None,
                    "m2_export_first": float(item.m2_export_first) if item.m2_export_first else None,
                    "pct_showroom": item.pct_showroom,
                    "pct_distribution": item.pct_distribution,
                }

                # Upsert (insert or update on conflict)
                self.db.table(self.table).upsert(
                    upsert_data,
                    on_conflict="schedule_date,production_date,factory_code,plant"
                ).execute()

                items_saved += 1

            except Exception as e:
                logger.error(
                    "save_schedule_item_failed",
                    factory_code=item.factory_code,
                    production_date=str(item.production_date),
                    error=str(e)
                )
                continue

        logger.info(
            "production_schedule_saved",
            items_saved=items_saved,
            products_matched=products_matched,
            unmatched_count=len(unmatched_codes)
        )

        return items_saved, products_matched, list(unmatched_codes)

    # ===================
    # READ OPERATIONS
    # ===================

    def get_by_schedule_date(
        self,
        schedule_date: date,
        plant: Optional[int] = None
    ) -> list[ProductionScheduleResponse]:
        """
        Get all production items for a specific schedule date.

        Args:
            schedule_date: The schedule generation date
            plant: Optional filter by plant (1 or 2)

        Returns:
            List of ProductionScheduleResponse
        """
        logger.debug("getting_schedule_by_date", schedule_date=str(schedule_date), plant=plant)

        try:
            query = (
                self.db.table(self.table)
                .select("*")
                .eq("schedule_date", schedule_date.isoformat())
                .order("production_date")
                .order("plant")
            )

            if plant:
                query = query.eq("plant", plant)

            result = query.execute()

            return [ProductionScheduleResponse(**row) for row in result.data]

        except Exception as e:
            logger.error("get_schedule_by_date_failed", error=str(e))
            raise DatabaseError("select", str(e))

    def get_upcoming_production(
        self,
        days_ahead: int = 30,
        product_id: Optional[str] = None,
        factory_code: Optional[str] = None
    ) -> list[UpcomingProductionItem]:
        """
        Get upcoming production within a date range.

        Args:
            days_ahead: Number of days to look ahead (default 30)
            product_id: Optional filter by product UUID
            factory_code: Optional filter by factory code

        Returns:
            List of UpcomingProductionItem sorted by date
        """
        logger.debug(
            "getting_upcoming_production",
            days_ahead=days_ahead,
            product_id=product_id,
            factory_code=factory_code
        )

        try:
            today = date.today()
            end_date = today + timedelta(days=days_ahead)

            # Use scheduled_start_date (actual DB column) instead of production_date
            query = (
                self.db.table(self.table)
                .select("*, products(sku)")
                .gte("scheduled_start_date", today.isoformat())
                .lte("scheduled_start_date", end_date.isoformat())
                .order("scheduled_start_date")
            )

            if product_id:
                query = query.eq("product_id", product_id)
            if factory_code:
                query = query.eq("factory_item_code", factory_code)

            result = query.execute()

            items = []
            for row in result.data:
                # Calculate days until production
                # Use scheduled_start_date as the production date
                prod_date_str = row.get("scheduled_start_date") or row.get("production_date")
                if not prod_date_str:
                    continue
                prod_date = date.fromisoformat(prod_date_str) if isinstance(prod_date_str, str) else prod_date_str
                days_until = (prod_date - today).days

                # Get SKU from joined products table or direct field
                sku = row.get("sku")
                if not sku and row.get("products"):
                    sku = row["products"].get("sku")

                items.append(UpcomingProductionItem(
                    production_date=prod_date,
                    factory_code=row.get("factory_item_code") or row.get("factory_code"),
                    product_name=row.get("referencia") or row.get("product_name"),
                    product_id=row.get("product_id"),
                    sku=sku,
                    plant=parse_plant_number(row.get("plant")),
                    m2_export_first=Decimal(str(row.get("requested_m2") or row.get("m2_export_first") or 0)),
                    days_until_production=days_until
                ))

            return items

        except Exception as e:
            logger.error("get_upcoming_production_failed", error=str(e))
            # Return empty list instead of raising to avoid breaking Order Builder
            return []

    def get_production_for_product(
        self,
        product_id: str,
        include_past: bool = False
    ) -> list[ProductionScheduleResponse]:
        """
        Get all production schedule entries for a specific product.

        Args:
            product_id: Product UUID
            include_past: If True, include past production dates

        Returns:
            List of ProductionScheduleResponse
        """
        logger.debug("getting_production_for_product", product_id=product_id)

        try:
            query = (
                self.db.table(self.table)
                .select("*")
                .eq("product_id", product_id)
                .order("production_date", desc=True)
            )

            if not include_past:
                query = query.gte("production_date", date.today().isoformat())

            result = query.execute()

            return [ProductionScheduleResponse(**row) for row in result.data]

        except Exception as e:
            logger.error("get_production_for_product_failed", error=str(e))
            raise DatabaseError("select", str(e))

    def get_unmatched_factory_codes(self) -> list[dict]:
        """
        Get factory codes that haven't been matched to products.

        Useful for identifying products that need factory_code set.

        Returns:
            List of dicts with factory_code, product_name, count
        """
        logger.debug("getting_unmatched_factory_codes")

        try:
            # Get distinct unmatched codes with their product names
            result = (
                self.db.table(self.table)
                .select("factory_code, product_name")
                .is_("product_id", "null")
                .execute()
            )

            # Aggregate by factory code
            codes = {}
            for row in result.data:
                code = row["factory_code"]
                if code not in codes:
                    codes[code] = {
                        "factory_code": code,
                        "product_name": row.get("product_name"),
                        "count": 0
                    }
                codes[code]["count"] += 1

            return list(codes.values())

        except Exception as e:
            logger.error("get_unmatched_factory_codes_failed", error=str(e))
            raise DatabaseError("select", str(e))

    def get_schedule_dates(self, limit: int = 10) -> list[dict]:
        """
        Get list of distinct schedule dates (most recent first).

        Returns:
            List of dicts with schedule_date, schedule_version, item_count
        """
        logger.debug("getting_schedule_dates", limit=limit)

        try:
            # Get all records and aggregate in Python (Supabase doesn't support GROUP BY well)
            result = (
                self.db.table(self.table)
                .select("schedule_date, schedule_version")
                .order("schedule_date", desc=True)
                .execute()
            )

            # Aggregate by schedule_date
            schedules = {}
            for row in result.data:
                sched_date = row["schedule_date"]
                if sched_date not in schedules:
                    schedules[sched_date] = {
                        "schedule_date": sched_date,
                        "schedule_version": row.get("schedule_version"),
                        "item_count": 0
                    }
                schedules[sched_date]["item_count"] += 1

            # Sort and limit
            sorted_schedules = sorted(
                schedules.values(),
                key=lambda x: x["schedule_date"],
                reverse=True
            )[:limit]

            return sorted_schedules

        except Exception as e:
            logger.error("get_schedule_dates_failed", error=str(e))
            raise DatabaseError("select", str(e))

    # ===================
    # UTILITY OPERATIONS
    # ===================

    def rematch_products(self) -> tuple[int, int]:
        """
        Re-match all schedule items to products by factory code.

        Useful after updating factory_code on products.

        Returns:
            Tuple of (total_processed, newly_matched)
        """
        logger.info("rematching_products")

        try:
            # Get all unmatched items
            unmatched_result = (
                self.db.table(self.table)
                .select("id, factory_code")
                .is_("product_id", "null")
                .execute()
            )

            if not unmatched_result.data:
                return 0, 0

            # Get unique factory codes
            factory_codes = list(set(row["factory_code"] for row in unmatched_result.data))
            products = self.product_service.get_by_factory_codes(factory_codes)
            products_by_code = {p.factory_code: p for p in products if p.factory_code}

            newly_matched = 0
            for row in unmatched_result.data:
                product = products_by_code.get(row["factory_code"])
                if product:
                    self.db.table(self.table).update(
                        {"product_id": product.id}
                    ).eq("id", row["id"]).execute()
                    newly_matched += 1

            logger.info(
                "products_rematched",
                total_processed=len(unmatched_result.data),
                newly_matched=newly_matched
            )

            return len(unmatched_result.data), newly_matched

        except Exception as e:
            logger.error("rematch_products_failed", error=str(e))
            raise DatabaseError("update", str(e))

    # ===================
    # ORDER BUILDER INTEGRATION
    # ===================

    def wipe_and_replace(
        self,
        parsed_data: ParsedProductionSchedule,
        filename: Optional[str] = None
    ) -> UploadResult:
        """
        Wipe existing production schedule and replace with new data.

        Used for daily PDF uploads that replace the entire schedule.

        Args:
            parsed_data: Parsed schedule from Claude Vision
            filename: Original PDF filename

        Returns:
            UploadResult with matched/unmatched counts
        """
        logger.info(
            "wiping_and_replacing_schedule",
            schedule_date=str(parsed_data.schedule_date),
            line_items_count=len(parsed_data.line_items),
            filename=filename
        )

        warnings = []

        try:
            # Step 1: Wipe existing data
            delete_result = (
                self.db.table(self.table)
                .delete()
                .neq("id", "00000000-0000-0000-0000-000000000000")  # Delete all
                .execute()
            )
            logger.info("existing_schedule_deleted")

        except Exception as e:
            warnings.append(f"Could not delete existing data: {str(e)}")
            logger.warning("delete_existing_failed", error=str(e))

        # Step 2: Insert new data
        items_saved, products_matched, unmatched_codes = self.save_parsed_schedule(
            parsed_data,
            filename
        )

        # Step 3: Build unmatched product details with fuzzy suggestions
        unmatched_products = self._get_unmatched_with_suggestions(parsed_data, unmatched_codes)

        return UploadResult(
            total_rows=items_saved,
            matched_count=products_matched,
            unmatched_count=len(unmatched_codes),
            schedule_date=parsed_data.schedule_date,
            schedule_version=parsed_data.schedule_version,
            filename=filename or "unknown.pdf",
            unmatched_products=unmatched_products,
            warnings=warnings,
        )

    def _get_unmatched_with_suggestions(
        self,
        parsed_data: ParsedProductionSchedule,
        unmatched_codes: list[str]
    ) -> list[UnmappedProduct]:
        """
        Build UnmappedProduct list with fuzzy match suggestions.

        Args:
            parsed_data: Parsed schedule data
            unmatched_codes: List of factory codes without product match

        Returns:
            List of UnmappedProduct with suggestions
        """
        if not unmatched_codes:
            return []

        # Group line items by factory code
        items_by_code: dict[str, list] = defaultdict(list)
        for item in parsed_data.line_items:
            if item.factory_code in unmatched_codes:
                items_by_code[item.factory_code].append(item)

        # Get all products for fuzzy matching
        all_products = []
        try:
            all_products = self.product_service.get_all_active_tiles()
        except Exception as e:
            logger.warning("get_products_for_fuzzy_failed", error=str(e))

        unmatched_products = []
        for factory_code in unmatched_codes:
            items = items_by_code.get(factory_code, [])
            if not items:
                continue

            # Get factory name from first item
            factory_name = items[0].product_name or factory_code

            # Sum total m²
            total_m2 = sum(
                item.m2_total_net or Decimal("0")
                for item in items
            )

            # Get unique production dates
            production_dates = sorted(set(
                item.production_date.isoformat()
                for item in items
            ))

            # Generate fuzzy match suggestions
            suggestions = self._get_fuzzy_suggestions(factory_name, all_products)

            unmatched_products.append(UnmappedProduct(
                factory_code=factory_code,
                factory_name=factory_name,
                total_m2=total_m2,
                production_dates=production_dates,
                row_count=len(items),
                suggested_matches=suggestions,
            ))

        return unmatched_products

    def _get_fuzzy_suggestions(
        self,
        factory_name: str,
        products: list,
        limit: int = 3
    ) -> list[MatchSuggestion]:
        """
        Generate fuzzy match suggestions for a factory product name.

        Args:
            factory_name: Product name from factory PDF
            products: List of product objects with id and sku
            limit: Max suggestions to return

        Returns:
            List of MatchSuggestion sorted by score
        """
        if not FUZZY_AVAILABLE or not products:
            return []

        suggestions = []
        for product in products:
            sku = getattr(product, "sku", None)
            if not sku:
                continue

            # Calculate fuzzy match score (convert to int for model)
            score = int(fuzz.token_sort_ratio(factory_name.upper(), sku.upper()))

            if score >= 50:  # Minimum threshold
                suggestions.append(MatchSuggestion(
                    product_id=product.id,
                    sku=sku,
                    score=score,
                    match_reason="fuzzy_name" if score >= 80 else "partial_match"
                ))

        # Sort by score descending and limit
        suggestions.sort(key=lambda s: s.score, reverse=True)
        return suggestions[:limit]

    def get_factory_status(
        self,
        product_ids: list[str],
        boat_departure: date,
        buffer_days: int = 3
    ) -> dict[str, ProductFactoryStatus]:
        """
        Get factory production status for Order Builder.

        Checks if products have scheduled production and whether
        they'll be ready before the boat departs.

        Args:
            product_ids: List of product UUIDs to check
            boat_departure: Boat departure date for timing assessment
            buffer_days: Days buffer before boat (default 3)

        Returns:
            Dict mapping product_id to ProductFactoryStatus
        """
        if not product_ids:
            return {}

        logger.debug(
            "getting_factory_status",
            product_count=len(product_ids),
            boat_departure=str(boat_departure)
        )

        today = date.today()
        cutoff_date = boat_departure - timedelta(days=buffer_days)

        try:
            # Get upcoming production for these products
            # Use actual DB column names: scheduled_start_date, requested_m2
            result = (
                self.db.table(self.table)
                .select("product_id, sku, scheduled_start_date, requested_m2, status, products(sku)")
                .in_("product_id", product_ids)
                .gte("scheduled_start_date", today.isoformat())
                .order("scheduled_start_date")
                .execute()
            )

            # Group by product_id (take earliest production date)
            production_by_product: dict[str, dict] = {}
            for row in result.data:
                pid = row["product_id"]
                if pid not in production_by_product:
                    production_by_product[pid] = row

            # Build status for each requested product
            status_map = {}
            for pid in product_ids:
                prod = production_by_product.get(pid)

                if prod:
                    # Use scheduled_start_date as production_date
                    prod_date_str = prod.get("scheduled_start_date")
                    if prod_date_str:
                        prod_date = date.fromisoformat(prod_date_str) if isinstance(prod_date_str, str) else prod_date_str
                        days_until = (prod_date - today).days
                        ready_before_boat = prod_date <= cutoff_date
                    else:
                        prod_date = None
                        days_until = None
                        ready_before_boat = False

                    m2 = Decimal(str(prod.get("requested_m2") or prod.get("m2_total_net") or 0))
                    sku = prod.get("sku") or ""
                    if not sku and prod.get("products"):
                        sku = prod["products"].get("sku", "")

                    # Build timing message
                    if prod_date:
                        if ready_before_boat:
                            timing_msg = f"Ready {prod_date.strftime('%b %d')} - in time"
                        else:
                            days_late = (prod_date - cutoff_date).days
                            timing_msg = f"Ready {prod_date.strftime('%b %d')} - {days_late}d after deadline"
                    else:
                        timing_msg = "Production scheduled"

                    status_map[pid] = ProductFactoryStatus(
                        product_id=pid,
                        sku=sku,
                        status=FactoryStatus.IN_PRODUCTION,
                        production_date=prod_date,
                        production_m2=m2,
                        days_until_ready=days_until,
                        ready_before_boat=ready_before_boat,
                        timing_message=timing_msg,
                    )
                else:
                    # No scheduled production
                    status_map[pid] = ProductFactoryStatus(
                        product_id=pid,
                        sku="",
                        status=FactoryStatus.NOT_SCHEDULED,
                        timing_message="Not in production schedule",
                    )

            logger.debug(
                "factory_status_retrieved",
                in_production=sum(1 for s in status_map.values() if s.status == FactoryStatus.IN_PRODUCTION),
                not_scheduled=sum(1 for s in status_map.values() if s.status == FactoryStatus.NOT_SCHEDULED),
            )

            return status_map

        except Exception as e:
            logger.error("get_factory_status_failed", error=str(e))
            # Return empty status for all products on error
            return {
                pid: ProductFactoryStatus(
                    product_id=pid,
                    sku="",
                    status=FactoryStatus.NOT_SCHEDULED,
                    timing_message="Status unavailable",
                )
                for pid in product_ids
            }

    def map_factory_code_to_product(
        self,
        factory_code: str,
        product_id: str
    ) -> int:
        """
        Map a factory code to a product and update all schedule rows.

        Also updates the product's factory_code field.

        Args:
            factory_code: Factory internal code
            product_id: Product UUID

        Returns:
            Number of schedule rows updated
        """
        logger.info(
            "mapping_factory_code",
            factory_code=factory_code,
            product_id=product_id
        )

        try:
            # Update product's factory_code
            self.db.table("products").update({
                "factory_code": factory_code
            }).eq("id", product_id).execute()

            # Update all schedule rows with this factory code
            result = (
                self.db.table(self.table)
                .update({"product_id": product_id})
                .eq("factory_code", factory_code)
                .execute()
            )

            rows_updated = len(result.data) if result.data else 0

            logger.info(
                "factory_code_mapped",
                factory_code=factory_code,
                product_id=product_id,
                rows_updated=rows_updated
            )

            return rows_updated

        except Exception as e:
            logger.error("map_factory_code_failed", error=str(e))
            raise DatabaseError("update", str(e))

    # ===================
    # EXCEL-BASED PARSING (Programa de Produccion)
    # ===================
    # These methods handle the Excel-based production schedule with color-coded status

    # Color constants for Excel parsing
    COLOR_GREEN = 'FF00B050'      # Completed
    COLOR_LIGHT_BLUE = 'FF00B0F0'  # In Progress
    COLOR_ORANGE = 'FFFFC000'      # Attention (treat as scheduled)

    def parse_production_excel(
        self,
        file_path: str,
        source_month: str = None
    ) -> list[ProductionScheduleCreate]:
        """
        Parse Programa de Produccion Excel file.

        Extracts Guatemala export data (m2 Primera exportacion) with status
        detection from cell colors:
        - Green = completed
        - Light Blue = in_progress
        - White/None = scheduled (CAN ADD MORE!)

        Args:
            file_path: Path to Excel file
            source_month: Month identifier (e.g., 'ENERO-26')

        Returns:
            List of ProductionScheduleCreate records
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel parsing")

        logger.info("parsing_production_excel", file_path=file_path)

        # Load workbook - need both data and formatting
        wb_data = load_workbook(file_path, data_only=True)
        wb_format = load_workbook(file_path, data_only=False)

        # Find the monthly sheet (e.g., "ENERO-26")
        sheet_name = None
        for name in wb_data.sheetnames:
            if "-26" in name or "-25" in name:  # Year suffix
                sheet_name = name
                break

        if not sheet_name:
            # Fallback to first sheet that looks like a month
            for name in wb_data.sheetnames:
                if any(month in name.upper() for month in [
                    'ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
                    'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE'
                ]):
                    sheet_name = name
                    break

        if not sheet_name:
            sheet_name = wb_data.sheetnames[1] if len(wb_data.sheetnames) > 1 else wb_data.sheetnames[0]

        ws_data = wb_data[sheet_name]
        ws_format = wb_format[sheet_name]

        if source_month is None:
            source_month = sheet_name

        logger.info("parsing_sheet", sheet_name=sheet_name)

        records = []
        import os
        filename = os.path.basename(file_path)

        # Parse Plant 1 (columns 5, 11, 17, 19)
        # Item Code = col 5, Referencia = col 11, Programa = col 17, Real = col 19
        # Dates: Fecha Inicio = col 1, Fecha Fin = col 2, Est Entrega = col 3
        plant1_records = self._parse_plant_data(
            ws_data, ws_format,
            plant="plant_1",
            item_col=5,
            ref_col=11,
            programa_col=17,
            real_col=19,
            start_row=18,
            end_row=50,
            source_file=filename,
            source_month=source_month,
            fecha_inicio_col=1,
            fecha_fin_col=2,
            fecha_entrega_col=3,
        )
        records.extend(plant1_records)

        # Parse Plant 2 (columns 20, 28, 34, 36)
        # Item Code = col 20, Referencia = col 28, Programa = col 34, Real = col 36
        # Dates: Fecha Inicio = col 23, Fecha Fin = col 24 (no delivery column for Plant 2)
        plant2_records = self._parse_plant_data(
            ws_data, ws_format,
            plant="plant_2",
            item_col=20,
            ref_col=28,
            programa_col=34,
            real_col=36,
            start_row=18,
            end_row=50,
            source_file=filename,
            source_month=source_month,
            fecha_inicio_col=23,
            fecha_fin_col=24,
            fecha_entrega_col=None,  # Plant 2 doesn't have delivery date column
        )
        records.extend(plant2_records)

        logger.info(
            "excel_parsed",
            total_records=len(records),
            plant1_count=len(plant1_records),
            plant2_count=len(plant2_records)
        )

        return records

    def _parse_plant_data(
        self,
        ws_data,
        ws_format,
        plant: str,
        item_col: int,
        ref_col: int,
        programa_col: int,
        real_col: int,
        start_row: int,
        end_row: int,
        source_file: str,
        source_month: str,
        fecha_inicio_col: int = None,
        fecha_fin_col: int = None,
        fecha_entrega_col: int = None,
    ) -> list[ProductionScheduleCreate]:
        """Parse data for one plant from Excel worksheet.

        Args:
            fecha_inicio_col: Column for Fecha Inicio (start date)
            fecha_fin_col: Column for Fecha Fin (end date)
            fecha_entrega_col: Column for Fecha estimada entrega (delivery date)
        """
        from datetime import datetime as dt

        records = []

        for row_num in range(start_row, end_row + 1):
            ref_value = ws_data.cell(row=row_num, column=ref_col).value
            programa_value = ws_data.cell(row=row_num, column=programa_col).value
            real_value = ws_data.cell(row=row_num, column=real_col).value

            # Skip rows without referencia or without Guatemala data
            if ref_value is None:
                continue

            # Convert to string for filtering
            ref_str = str(ref_value).strip()

            # Skip invalid references
            if '=' in ref_str:  # Formula
                continue
            if ref_str in ['MANTENIMIENTO DE PLANTA', 'REPROCESO', 'TOTAL', 'SUBTOTAL']:
                continue
            if ref_str.isdigit():  # Pure numbers (likely formula results like row counts)
                continue
            if len(ref_str) < 3:  # Too short to be a valid product name
                continue

            # Only process rows with Guatemala export data
            if programa_value is None and real_value is None:
                continue

            # Get item code
            item_code = ws_data.cell(row=row_num, column=item_col).value
            item_code_str = str(item_code) if item_code else None

            # Detect status from cell color
            status = self._get_status_from_color(
                ws_format.cell(row=row_num, column=ref_col)
            )

            # Parse m² values
            programa_m2 = Decimal(str(programa_value)) if programa_value else Decimal("0")
            real_m2 = Decimal(str(real_value)) if real_value else Decimal("0")

            # Parse date columns
            scheduled_start_date = None
            scheduled_end_date = None
            estimated_delivery_date = None

            if fecha_inicio_col:
                val = ws_data.cell(row=row_num, column=fecha_inicio_col).value
                if isinstance(val, dt):
                    scheduled_start_date = val.date()
                elif isinstance(val, date):
                    scheduled_start_date = val

            if fecha_fin_col:
                val = ws_data.cell(row=row_num, column=fecha_fin_col).value
                if isinstance(val, dt):
                    scheduled_end_date = val.date()
                elif isinstance(val, date):
                    scheduled_end_date = val

            if fecha_entrega_col:
                val = ws_data.cell(row=row_num, column=fecha_entrega_col).value
                if isinstance(val, dt):
                    estimated_delivery_date = val.date()
                elif isinstance(val, date):
                    estimated_delivery_date = val

            records.append(ProductionScheduleCreate(
                factory_item_code=item_code_str,
                referencia=str(ref_value).strip(),
                plant=plant,
                requested_m2=programa_m2,
                completed_m2=real_m2,
                status=status,
                scheduled_start_date=scheduled_start_date,
                scheduled_end_date=scheduled_end_date,
                estimated_delivery_date=estimated_delivery_date,
                source_file=source_file,
                source_month=source_month,
                source_row=row_num,
            ))

        return records

    def _get_status_from_color(self, cell) -> ProductionStatus:
        """Determine production status from cell background color."""
        fill = cell.fill
        if fill.fill_type is None or fill.fill_type == 'none':
            return ProductionStatus.SCHEDULED  # White = CAN ADD MORE!

        fg = fill.fgColor
        if fg is None:
            return ProductionStatus.SCHEDULED

        if fg.type == 'rgb' and fg.rgb:
            rgb = fg.rgb
            if rgb == self.COLOR_GREEN:
                return ProductionStatus.COMPLETED
            elif rgb == self.COLOR_LIGHT_BLUE:
                return ProductionStatus.IN_PROGRESS
            elif rgb == self.COLOR_ORANGE:
                return ProductionStatus.SCHEDULED  # Orange = needs attention but not started

        return ProductionStatus.SCHEDULED

    def import_from_excel(
        self,
        records: list[ProductionScheduleCreate],
        match_products: bool = True
    ) -> ProductionImportResult:
        """
        Import parsed Excel records to database.

        IMPORTANT: Same SKU can appear multiple times in Excel (different production runs).
        This method SUMS requested_m2 and completed_m2 for records with same referencia+plant.

        Args:
            records: List of ProductionScheduleCreate from parse_production_excel()
            match_products: Whether to attempt SKU matching

        Returns:
            ProductionImportResult with counts and warnings
        """
        if not records:
            return ProductionImportResult(
                filename="",
                source_month="",
                total_rows_parsed=0,
                rows_with_guatemala_data=0,
                warnings=["No records to import"]
            )

        logger.info("importing_excel_records", count=len(records))

        # Get source info from first record
        filename = records[0].source_file or ""
        source_month = records[0].source_month or ""

        # =====================================================
        # STEP 1: Aggregate records by (referencia, plant)
        # Same SKU in same plant should be SUMMED, not overwritten
        # =====================================================
        aggregated: dict[tuple, dict] = {}
        status_priority = {
            ProductionStatus.COMPLETED: 3,
            ProductionStatus.IN_PROGRESS: 2,
            ProductionStatus.SCHEDULED: 1,
        }

        for record in records:
            key = (record.referencia, record.plant)

            if key not in aggregated:
                # First occurrence - initialize
                aggregated[key] = {
                    "record": record,
                    "requested_m2": record.requested_m2,
                    "completed_m2": record.completed_m2,
                    "status": record.status,
                    "row_count": 1,
                }
            else:
                # Same SKU+plant - SUM the values
                agg = aggregated[key]
                agg["requested_m2"] += record.requested_m2
                agg["completed_m2"] += record.completed_m2
                agg["row_count"] += 1

                # Keep the more significant status (completed > in_progress > scheduled)
                if status_priority.get(record.status, 0) > status_priority.get(agg["status"], 0):
                    agg["status"] = record.status

                # Keep dates from completed row if available
                if record.status == ProductionStatus.COMPLETED:
                    agg["record"] = record

        logger.info(
            "records_aggregated",
            raw_count=len(records),
            aggregated_count=len(aggregated),
            duplicates_merged=len(records) - len(aggregated)
        )

        # Track results
        inserted = 0
        updated = 0
        skipped = 0
        matched = 0
        unmatched_refs = set()
        warnings = []

        # Status counts (from aggregated data)
        status_counts = {
            ProductionStatus.COMPLETED: 0,
            ProductionStatus.IN_PROGRESS: 0,
            ProductionStatus.SCHEDULED: 0,
        }
        total_requested = Decimal("0")
        total_completed = Decimal("0")

        # Get products for matching
        products_by_sku = {}
        if match_products:
            try:
                all_products = self.product_service.get_all_active_tiles()
                for p in all_products:
                    # Index by SKU and normalized forms (accent-normalized)
                    sku = getattr(p, 'sku', None)
                    if sku:
                        sku_upper = sku.upper()
                        sku_no_accent = normalize_accents(sku_upper)
                        # Index by original SKU
                        products_by_sku[sku_upper] = p
                        # Index by accent-normalized SKU
                        products_by_sku[sku_no_accent] = p
                        # Also index without BTE suffix
                        no_bte = sku_upper.replace(' BTE', '').replace('BTE', '').strip()
                        no_bte_no_accent = normalize_accents(no_bte)
                        products_by_sku[no_bte] = p
                        products_by_sku[no_bte_no_accent] = p
            except Exception as e:
                warnings.append(f"Could not load products for matching: {e}")

        # =====================================================
        # STEP 2: Process aggregated records
        # =====================================================
        for (referencia, plant), agg in aggregated.items():
            record = agg["record"]
            # Use AGGREGATED values (summed from all rows with same SKU+plant)
            agg_requested = agg["requested_m2"]
            agg_completed = agg["completed_m2"]
            agg_status = agg["status"]

            try:
                # Attempt to match product
                product_id = None
                sku = None

                if match_products and referencia:
                    ref_upper = referencia.upper()
                    ref_no_accent = normalize_accents(ref_upper)
                    ref_no_bte = ref_upper.replace(' BTE', '').replace('BTE', '').strip()
                    ref_no_bte_no_accent = normalize_accents(ref_no_bte)

                    # Try matching in order: exact, no accent, no BTE, no accent + no BTE
                    product = (
                        products_by_sku.get(ref_upper) or
                        products_by_sku.get(ref_no_accent) or
                        products_by_sku.get(ref_no_bte) or
                        products_by_sku.get(ref_no_bte_no_accent)
                    )
                    if product:
                        product_id = product.id
                        sku = getattr(product, 'sku', None)
                        matched += 1
                    else:
                        unmatched_refs.add(referencia)

                # Build upsert data with AGGREGATED values
                upsert_data = {
                    "factory_item_code": record.factory_item_code,
                    "referencia": referencia,
                    "sku": sku,
                    "product_id": product_id,
                    "plant": plant,
                    "requested_m2": float(agg_requested),  # SUMMED value
                    "completed_m2": float(agg_completed),  # SUMMED value
                    "status": agg_status.value,            # Most significant status
                    "scheduled_start_date": record.scheduled_start_date.isoformat() if record.scheduled_start_date else None,
                    "scheduled_end_date": record.scheduled_end_date.isoformat() if record.scheduled_end_date else None,
                    "estimated_delivery_date": record.estimated_delivery_date.isoformat() if record.estimated_delivery_date else None,
                    "source_file": record.source_file,
                    "source_month": record.source_month,
                    "source_row": record.source_row,
                }

                # Upsert (use referencia + plant + source_month as unique key)
                result = self.db.table(self.table).upsert(
                    upsert_data,
                    on_conflict="referencia,plant,source_month"
                ).execute()

                if result.data:
                    inserted += 1  # Could be insert or update

                # Update totals with AGGREGATED values
                status_counts[agg_status] += 1
                total_requested += agg_requested
                total_completed += agg_completed

                # Log if rows were merged
                if agg["row_count"] > 1:
                    logger.info(
                        "sku_rows_merged",
                        referencia=referencia,
                        plant=plant,
                        row_count=agg["row_count"],
                        total_requested=float(agg_requested),
                        total_completed=float(agg_completed)
                    )

            except Exception as e:
                logger.warning(
                    "import_record_failed",
                    referencia=referencia,
                    error=str(e)
                )
                skipped += 1
                warnings.append(f"Failed to import {referencia}: {e}")

        logger.info(
            "excel_import_complete",
            inserted=inserted,
            matched=matched,
            unmatched=len(unmatched_refs),
            skipped=skipped
        )

        # Save history snapshot for slip tracking
        try:
            history_saved = self.save_history_snapshot(records, source_file=filename)
            logger.info("history_snapshot_saved_on_import", records_saved=history_saved)
        except Exception as e:
            warnings.append(f"Failed to save history snapshot: {e}")
            logger.warning("history_snapshot_save_failed", error=str(e))

        return ProductionImportResult(
            filename=filename,
            source_month=source_month,
            total_rows_parsed=len(records),
            rows_with_guatemala_data=len(records),
            inserted=inserted,
            updated=updated,
            skipped=skipped,
            matched_to_products=matched,
            unmatched_referencias=list(unmatched_refs),
            completed_count=status_counts[ProductionStatus.COMPLETED],
            in_progress_count=status_counts[ProductionStatus.IN_PROGRESS],
            scheduled_count=status_counts[ProductionStatus.SCHEDULED],
            total_requested_m2=total_requested,
            total_completed_m2=total_completed,
            warnings=warnings,
        )

    def get_production_summary(self) -> list[ProductionSummary]:
        """
        Get production summary by status.

        Returns breakdown of completed, in_progress, and scheduled items.
        """
        logger.debug("getting_production_summary")

        try:
            result = self.db.table(self.table).select(
                "status, requested_m2, completed_m2"
            ).execute()

            # Aggregate by status
            summaries = {
                "completed": {"count": 0, "requested": Decimal("0"), "completed": Decimal("0")},
                "in_progress": {"count": 0, "requested": Decimal("0"), "completed": Decimal("0")},
                "scheduled": {"count": 0, "requested": Decimal("0"), "completed": Decimal("0")},
            }

            for row in result.data:
                status = row.get("status", "scheduled")
                if status in summaries:
                    summaries[status]["count"] += 1
                    summaries[status]["requested"] += Decimal(str(row.get("requested_m2", 0)))
                    summaries[status]["completed"] += Decimal(str(row.get("completed_m2", 0)))

            action_hints = {
                "completed": "READY TO SHIP",
                "in_progress": "MANUFACTURING",
                "scheduled": "CAN ADD MORE",
            }

            return [
                ProductionSummary(
                    status=ProductionStatus(status),
                    item_count=data["count"],
                    total_requested_m2=data["requested"],
                    total_completed_m2=data["completed"],
                    total_remaining_m2=data["requested"] - data["completed"],
                    action_hint=action_hints[status],
                )
                for status, data in summaries.items()
                if data["count"] > 0
            ]

        except Exception as e:
            logger.error("get_production_summary_failed", error=str(e))
            return []

    def get_can_add_more_items(self) -> list[ProductionScheduleDBResponse]:
        """
        Get items in 'scheduled' status that can have more quantity added.

        These are items where production hasn't started yet.

        Returns:
            List of ProductionScheduleDBResponse with can_add_more=True
        """
        logger.debug("getting_can_add_more_items")

        try:
            result = (
                self.db.table(self.table)
                .select("*")
                .eq("status", "scheduled")
                .order("referencia")
                .execute()
            )

            return [
                ProductionScheduleDBResponse.from_db(row)
                for row in result.data
            ]

        except Exception as e:
            logger.error("get_can_add_more_items_failed", error=str(e))
            return []

    def get_capacity(self, monthly_limit_m2: Decimal = Decimal("60000")) -> ProductionCapacity:
        """
        Get factory request capacity tracking.

        Args:
            monthly_limit_m2: Monthly quota (default 60,000 m² for Guatemala)

        Returns:
            ProductionCapacity with utilization breakdown
        """
        logger.debug("getting_production_capacity")

        try:
            result = self.db.table(self.table).select(
                "status, requested_m2, referencia"
            ).execute()

            completed_m2 = Decimal("0")
            in_progress_m2 = Decimal("0")
            scheduled_m2 = Decimal("0")
            can_add_items = []

            for row in result.data:
                requested = Decimal(str(row.get("requested_m2", 0)))
                status = row.get("status", "scheduled")

                if status == "completed":
                    completed_m2 += requested
                elif status == "in_progress":
                    in_progress_m2 += requested
                elif status == "scheduled":
                    scheduled_m2 += requested
                    can_add_items.append(row.get("referencia", "Unknown"))

            total_requested = completed_m2 + in_progress_m2 + scheduled_m2
            available = max(Decimal("0"), monthly_limit_m2 - total_requested)
            utilization = (total_requested / monthly_limit_m2 * 100) if monthly_limit_m2 > 0 else Decimal("0")

            return ProductionCapacity(
                monthly_limit_m2=monthly_limit_m2,
                already_requested_m2=total_requested,
                available_to_request_m2=available,
                utilization_pct=utilization,
                completed_m2=completed_m2,
                in_progress_m2=in_progress_m2,
                scheduled_m2=scheduled_m2,
                can_add_more_items=can_add_items,
            )

        except Exception as e:
            logger.error("get_production_capacity_failed", error=str(e))
            return ProductionCapacity()

    def get_production_by_sku(self) -> dict[str, ProductionScheduleDBResponse]:
        """
        Get production status for all products, indexed by SKU.

        Used by Order Builder to enrich product data with production status.
        For products with multiple production entries (e.g., in different plants),
        returns the most recent entry.

        Returns:
            Dict mapping SKU to ProductionScheduleDBResponse
        """
        logger.debug("getting_production_by_sku")

        try:
            result = (
                self.db.table(self.table)
                .select("*")
                .not_.is_("sku", "null")
                .order("updated_at", desc=True)
                .execute()
            )

            # Build map, taking first (most recent) entry per SKU
            production_map: dict[str, ProductionScheduleDBResponse] = {}
            for row in result.data:
                sku = row.get("sku")
                if sku and sku not in production_map:
                    production_map[sku] = ProductionScheduleDBResponse.from_db(row)

            logger.info(
                "production_by_sku_retrieved",
                total_records=len(result.data),
                unique_skus=len(production_map)
            )

            return production_map

        except Exception as e:
            logger.error("get_production_by_sku_failed", error=str(e))
            return {}

    def get_production_for_order_builder(
        self,
        skus: list[str]
    ) -> dict[str, ProductionScheduleDBResponse]:
        """
        Get production status for specific SKUs (batch lookup).

        More efficient than get_production_by_sku when you only need
        production data for specific products.

        Args:
            skus: List of SKUs to look up

        Returns:
            Dict mapping SKU to ProductionScheduleDBResponse
        """
        if not skus:
            return {}

        logger.debug("getting_production_for_order_builder", sku_count=len(skus))

        try:
            # Normalize SKUs for matching
            normalized_skus = [normalize_accents(sku.upper()) for sku in skus]

            result = (
                self.db.table(self.table)
                .select("*")
                .not_.is_("sku", "null")
                .order("updated_at", desc=True)
                .execute()
            )

            # Build map with accent-normalized lookup
            production_map: dict[str, ProductionScheduleDBResponse] = {}
            for row in result.data:
                sku = row.get("sku")
                if sku:
                    sku_normalized = normalize_accents(sku.upper())
                    # Check if this SKU matches any requested SKU
                    for original_sku, norm_sku in zip(skus, normalized_skus):
                        if sku_normalized == norm_sku and original_sku not in production_map:
                            production_map[original_sku] = ProductionScheduleDBResponse.from_db(row)
                            break

            logger.info(
                "production_for_order_builder_retrieved",
                requested=len(skus),
                found=len(production_map)
            )

            return production_map

        except Exception as e:
            logger.error("get_production_for_order_builder_failed", error=str(e))
            return {}

    def get_average_production_time(self, fallback_days: int = 7) -> int:
        """
        Calculate average production time from completed items.

        Uses actual_completion_date - scheduled_start_date for completed items.
        Falls back to estimated_delivery_date - scheduled_start_date if actual not available.

        Args:
            fallback_days: Default if no data available (default 7 days)

        Returns:
            Average production time in days
        """
        logger.debug("calculating_average_production_time")

        try:
            result = (
                self.db.table(self.table)
                .select("scheduled_start_date, scheduled_end_date, estimated_delivery_date, actual_completion_date, status")
                .eq("status", "completed")
                .not_.is_("scheduled_start_date", "null")
                .execute()
            )

            if not result.data:
                logger.info("no_completed_production_data", fallback=fallback_days)
                return fallback_days

            total_days = 0
            valid_count = 0

            for row in result.data:
                start_date = row.get("scheduled_start_date")
                if not start_date:
                    continue

                # Parse start date
                if isinstance(start_date, str):
                    start_date = date.fromisoformat(start_date)

                # Try actual completion date first, then estimated delivery, then scheduled end
                end_date = None
                for field in ["actual_completion_date", "estimated_delivery_date", "scheduled_end_date"]:
                    val = row.get(field)
                    if val:
                        end_date = date.fromisoformat(val) if isinstance(val, str) else val
                        break

                if end_date and end_date >= start_date:
                    days = (end_date - start_date).days
                    if days > 0 and days < 90:  # Sanity check: 1-90 days
                        total_days += days
                        valid_count += 1

            if valid_count > 0:
                avg_days = total_days // valid_count
                logger.info(
                    "average_production_time_calculated",
                    avg_days=avg_days,
                    sample_size=valid_count
                )
                return max(1, avg_days)  # At least 1 day

            logger.info("no_valid_production_time_data", fallback=fallback_days)
            return fallback_days

        except Exception as e:
            logger.error("get_average_production_time_failed", error=str(e))
            return fallback_days


    # ===================
    # HISTORY TRACKING
    # ===================

    def save_history_snapshot(
        self,
        records: list[ProductionScheduleCreate],
        source_file: Optional[str] = None
    ) -> int:
        """
        Save a snapshot of production schedule to history table.

        Called during Excel import to track schedule changes over time.
        Enables calculation of delivery slip statistics.

        Args:
            records: List of ProductionScheduleCreate from parse_production_excel()
            source_file: Source filename for tracking

        Returns:
            Number of history records saved
        """
        today = date.today()
        saved = 0

        for record in records:
            try:
                # Try to find product_id from SKU
                product_id = None
                if record.referencia:
                    try:
                        products = self.product_service.get_all_active_tiles()
                        ref_normalized = normalize_accents(record.referencia.upper())
                        for p in products:
                            sku = getattr(p, 'sku', '')
                            if normalize_accents(sku.upper()) == ref_normalized:
                                product_id = p.id
                                break
                    except Exception:
                        pass

                upsert_data = {
                    "sku": record.referencia,
                    "product_id": product_id,
                    "factory_item_code": record.factory_item_code,
                    "snapshot_date": today.isoformat(),
                    "source_file": source_file,
                    "scheduled_start_date": record.scheduled_start_date.isoformat() if record.scheduled_start_date else None,
                    "scheduled_end_date": record.scheduled_end_date.isoformat() if record.scheduled_end_date else None,
                    "estimated_delivery_date": record.estimated_delivery_date.isoformat() if record.estimated_delivery_date else None,
                    "status": record.status.value,
                    "requested_m2": float(record.requested_m2),
                    "completed_m2": float(record.completed_m2),
                }

                self.db.table("production_schedule_history").upsert(
                    upsert_data,
                    on_conflict="sku,snapshot_date"
                ).execute()
                saved += 1

            except Exception as e:
                logger.warning(
                    "save_history_snapshot_failed",
                    sku=record.referencia,
                    error=str(e)
                )

        logger.info(
            "history_snapshot_saved",
            snapshot_date=str(today),
            records_saved=saved
        )

        return saved

    def record_completion(
        self,
        sku: str,
        actual_date: date
    ) -> Optional[int]:
        """
        Record actual completion date and calculate slip.

        Called when production status changes to 'completed'.

        Args:
            sku: Product SKU (referencia)
            actual_date: Actual completion date

        Returns:
            Slip days (positive = late, negative = early), or None if no estimate found
        """
        logger.debug("recording_completion", sku=sku, actual_date=str(actual_date))

        try:
            # Find the most recent estimate for this SKU
            result = (
                self.db.table("production_schedule_history")
                .select("id, estimated_delivery_date")
                .eq("sku", sku)
                .not_.is_("estimated_delivery_date", "null")
                .order("snapshot_date", desc=True)
                .limit(1)
                .execute()
            )

            if not result.data:
                logger.debug("no_estimate_found_for_completion", sku=sku)
                return None

            row = result.data[0]
            estimated = date.fromisoformat(row["estimated_delivery_date"])
            slip_days = (actual_date - estimated).days

            # Update the history record
            self.db.table("production_schedule_history").update({
                "actual_completion_date": actual_date.isoformat(),
                "slip_days": slip_days,
                "status": "completed"
            }).eq("id", row["id"]).execute()

            logger.info(
                "completion_recorded",
                sku=sku,
                estimated=str(estimated),
                actual=str(actual_date),
                slip_days=slip_days
            )

            return slip_days

        except Exception as e:
            logger.error("record_completion_failed", sku=sku, error=str(e))
            return None

    def get_slip_statistics(self) -> dict:
        """
        Calculate slip statistics from historical data.

        Used to potentially adjust production_buffer_days dynamically.

        Returns:
            Dict with slip statistics:
            - total_completed: Number of items with slip data
            - avg_slip_days: Average slip (positive = late)
            - p90_slip_days: 90th percentile slip
            - max_slip_days: Maximum slip
            - min_slip_days: Minimum slip (negative = early)
            - recommended_buffer: Suggested buffer based on P90
        """
        logger.debug("calculating_slip_statistics")

        try:
            result = (
                self.db.table("production_schedule_history")
                .select("slip_days")
                .not_.is_("slip_days", "null")
                .execute()
            )

            if not result.data:
                return {
                    "total_completed": 0,
                    "avg_slip_days": 0,
                    "p90_slip_days": 0,
                    "max_slip_days": 0,
                    "min_slip_days": 0,
                    "recommended_buffer": settings.production_buffer_days,
                    "note": "No historical slip data available yet"
                }

            slip_values = [row["slip_days"] for row in result.data]
            slip_values.sort()

            total = len(slip_values)
            avg_slip = sum(slip_values) / total
            p90_index = int(total * 0.9)
            p90_slip = slip_values[min(p90_index, total - 1)]

            # Recommended buffer: at least 5 days, or P90 slip rounded up
            import math
            recommended = max(5, math.ceil(p90_slip))

            stats = {
                "total_completed": total,
                "avg_slip_days": round(avg_slip, 1),
                "p90_slip_days": p90_slip,
                "max_slip_days": max(slip_values),
                "min_slip_days": min(slip_values),
                "recommended_buffer": recommended
            }

            logger.info("slip_statistics_calculated", **stats)
            return stats

        except Exception as e:
            logger.error("get_slip_statistics_failed", error=str(e))
            return {
                "total_completed": 0,
                "avg_slip_days": 0,
                "p90_slip_days": 0,
                "max_slip_days": 0,
                "min_slip_days": 0,
                "recommended_buffer": settings.production_buffer_days,
                "error": str(e)
            }


    # ===================
    # ORDER BUILDER INTEGRATION
    # ===================

    def create_from_order_builder(
        self,
        items: list[dict],
        boat_departure: Optional[str] = None,
    ) -> list[dict]:
        """
        Create production_schedule rows from Order Builder factory request export.

        Closes the feedback loop: export Section 3 → INSERT production_schedule
        → recommendation_service reads it back → gap closes.

        Args:
            items: List of dicts with product_id, sku, requested_m2, referencia
            boat_departure: Boat departure date string (for estimated_delivery_date calc)

        Returns:
            List of created rows
        """
        if not items:
            return []

        logger.info(
            "creating_production_from_ob",
            item_count=len(items),
            boat_departure=boat_departure,
        )

        rows = []
        for item in items:
            row = {
                "product_id": item["product_id"],
                "sku": item.get("sku"),
                "referencia": item.get("referencia") or item.get("sku") or "ORDER_BUILDER",
                "plant": "plant_1",
                "requested_m2": float(item["requested_m2"]),
                "completed_m2": 0,
                "status": "scheduled",
                "source_file": "ORDER_BUILDER",
                "source_month": boat_departure[:7] if boat_departure else None,
            }

            # estimated_delivery_date = boat_departure - transport_to_port_days
            if boat_departure:
                from datetime import date as date_type, timedelta
                try:
                    dep = date_type.fromisoformat(boat_departure)
                    # Factory needs ~5 days to transport to port
                    row["estimated_delivery_date"] = (dep - timedelta(days=5)).isoformat()
                except ValueError:
                    pass

            rows.append(row)

        try:
            result = self.db.table(self.table).insert(rows).execute()
            logger.info(
                "production_from_ob_created",
                created=len(result.data),
            )
            return result.data
        except Exception as e:
            logger.error("production_from_ob_failed", error=str(e))
            raise DatabaseError("insert", str(e))

    def update_piggyback(
        self,
        items: list[dict],
    ) -> int:
        """
        Update production_schedule.requested_m2 for piggyback exports (Section 2).

        Closes the piggyback feedback loop: additional m2 added to existing
        production run → headroom shrinks → recommendation adjusts.

        Args:
            items: List of dicts with product_id, additional_m2

        Returns:
            Number of rows updated
        """
        if not items:
            return 0

        logger.info("updating_piggyback", item_count=len(items))

        updated = 0
        for item in items:
            try:
                # Find the scheduled production row for this product
                existing = (
                    self.db.table(self.table)
                    .select("id, requested_m2")
                    .eq("product_id", item["product_id"])
                    .eq("status", "scheduled")
                    .order("created_at", desc=True)
                    .limit(1)
                    .execute()
                )

                if existing.data:
                    row = existing.data[0]
                    new_requested = float(row["requested_m2"]) + float(item["additional_m2"])
                    self.db.table(self.table).update({
                        "requested_m2": new_requested,
                    }).eq("id", row["id"]).execute()
                    updated += 1
                    logger.debug(
                        "piggyback_updated",
                        product_id=item["product_id"],
                        old_m2=row["requested_m2"],
                        added_m2=item["additional_m2"],
                        new_m2=new_requested,
                    )
            except Exception as e:
                logger.error(
                    "piggyback_update_failed",
                    product_id=item["product_id"],
                    error=str(e),
                )

        logger.info("piggyback_complete", updated=updated)
        return updated


# Singleton instance
_schedule_service: Optional[ProductionScheduleService] = None


def get_production_schedule_service() -> ProductionScheduleService:
    """Get or create ProductionScheduleService instance."""
    global _schedule_service
    if _schedule_service is None:
        _schedule_service = ProductionScheduleService()
    return _schedule_service
