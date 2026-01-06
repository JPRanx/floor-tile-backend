"""
Tests for export_service — Factory order Excel generation.
"""

from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from services.export_service import (
    normalize_sku_for_factory,
    calculate_m2,
    calculate_containers,
    get_production_month,
    ExportService,
    get_export_service,
    M2_PER_PALLET_FACTORY,
    PALLETS_PER_CONTAINER,
)


class TestNormalizeSku:
    """Tests for SKU normalization."""

    def test_removes_bte_suffix(self):
        assert normalize_sku_for_factory("ALMENDRO BEIGE BTE") == "ALMENDRO BEIGE"

    def test_removes_bte_case_insensitive(self):
        assert normalize_sku_for_factory("ALMENDRO BEIGE bte") == "ALMENDRO BEIGE"

    def test_removes_t_marker(self):
        assert normalize_sku_for_factory("ALMENDRO BEIGE (T)") == "ALMENDRO BEIGE"

    def test_removes_format_suffix(self):
        assert normalize_sku_for_factory("ALMENDRO BEIGE 51X51") == "ALMENDRO BEIGE"

    def test_removes_format_with_number(self):
        assert normalize_sku_for_factory("ALMENDRO BEIGE 51X51-1") == "ALMENDRO BEIGE"

    def test_removes_all_combined(self):
        result = normalize_sku_for_factory("ALMENDRO BEIGE BTE (T) 51X51-1")
        assert result == "ALMENDRO BEIGE"

    def test_preserves_accents(self):
        assert normalize_sku_for_factory("NOGAL CAFE BTE") == "NOGAL CAFE"

    def test_preserves_special_characters(self):
        assert normalize_sku_for_factory("CEIBA GRIS CLARO BTE") == "CEIBA GRIS CLARO"

    def test_removes_trailing_numbers(self):
        assert normalize_sku_for_factory("ALMENDRO BEIGE-1") == "ALMENDRO BEIGE"
        assert normalize_sku_for_factory("ALMENDRO BEIGE-2") == "ALMENDRO BEIGE"

    def test_handles_empty_string(self):
        assert normalize_sku_for_factory("") == ""

    def test_handles_none(self):
        # Should not raise, handled gracefully
        assert normalize_sku_for_factory(None) == ""

    def test_strips_whitespace(self):
        assert normalize_sku_for_factory("  ALMENDRO BEIGE BTE  ") == "ALMENDRO BEIGE"


class TestCalculateM2:
    """Tests for m2 calculation."""

    def test_single_pallet(self):
        result = calculate_m2(1)
        assert result == Decimal("134.4")

    def test_fourteen_pallets(self):
        # 14 pallets = 1 container
        result = calculate_m2(14)
        assert result == Decimal("1881.6")

    def test_zero_pallets(self):
        result = calculate_m2(0)
        assert result == Decimal("0")

    def test_large_order(self):
        # 182 pallets = 13 containers worth
        result = calculate_m2(182)
        expected = Decimal("182") * M2_PER_PALLET_FACTORY
        assert result == expected


class TestCalculateContainers:
    """Tests for container calculation."""

    def test_exactly_one_container(self):
        # 14 pallets = exactly 1 container
        assert calculate_containers(14) == 1

    def test_partial_container_rounds_up(self):
        # 15 pallets = 2 containers (ceil)
        assert calculate_containers(15) == 2

    def test_zero_pallets(self):
        assert calculate_containers(0) == 0

    def test_negative_pallets(self):
        assert calculate_containers(-5) == 0

    def test_large_order(self):
        # 182 pallets / 14 = 13 containers
        assert calculate_containers(182) == 13

    def test_almost_full_container(self):
        # 13 pallets = still 1 container
        assert calculate_containers(13) == 1


class TestGetProductionMonth:
    """Tests for production month calculation."""

    def test_january_departure_gives_february(self):
        result = get_production_month(date(2026, 1, 15))
        assert result == "FEBRERO"

    def test_december_departure_gives_january(self):
        result = get_production_month(date(2025, 12, 20))
        assert result == "ENERO"

    def test_june_departure_gives_july(self):
        result = get_production_month(date(2026, 6, 10))
        assert result == "JULIO"

    def test_november_departure_gives_december(self):
        result = get_production_month(date(2026, 11, 5))
        assert result == "DICIEMBRE"


class TestExportService:
    """Tests for ExportService Excel generation."""

    @pytest.fixture
    def service(self):
        return ExportService()

    @pytest.fixture
    def sample_products(self):
        return [
            {"sku": "ALMENDRO BEIGE BTE (T) 51X51-1", "pallets": 14},
            {"sku": "CEIBA GRIS CLARO BTE", "pallets": 7},
            {"sku": "NOGAL CAFE BTE", "pallets": 14},
        ]

    def test_generates_valid_excel(self, service, sample_products):
        result = service.generate_factory_order_excel(
            products=sample_products,
            boat_departure=date(2026, 1, 15),
        )
        assert isinstance(result, BytesIO)

        # Verify it's a valid Excel file
        wb = load_workbook(result)
        assert "PEDIDO TARRAGONA" in wb.sheetnames

    def test_excel_has_correct_header(self, service, sample_products):
        result = service.generate_factory_order_excel(
            products=sample_products,
            boat_departure=date(2026, 1, 15),
        )
        wb = load_workbook(result)
        ws = wb.active

        assert ws["A1"].value == "Pedido Tarragona Guatemala"

    def test_excel_has_order_date(self, service, sample_products):
        order_date = date(2026, 1, 6)
        result = service.generate_factory_order_excel(
            products=sample_products,
            boat_departure=date(2026, 1, 15),
            order_date=order_date,
        )
        wb = load_workbook(result)
        ws = wb.active

        assert ws["A3"].value == "Fecha de pedido:"
        assert ws["B3"].value == "06/01/2026"

    def test_excel_has_production_month(self, service, sample_products):
        result = service.generate_factory_order_excel(
            products=sample_products,
            boat_departure=date(2026, 1, 15),
        )
        wb = load_workbook(result)
        ws = wb.active

        assert ws["A5"].value == "Fabricacion para:"
        assert ws["B5"].value == "FEBRERO"

    def test_excel_has_column_headers(self, service, sample_products):
        result = service.generate_factory_order_excel(
            products=sample_products,
            boat_departure=date(2026, 1, 15),
        )
        wb = load_workbook(result)
        ws = wb.active

        assert ws["A7"].value == "Referencia"
        assert ws["B7"].value == "Formato"
        assert ws["C7"].value == "M2 solicitados"

    def test_excel_has_normalized_skus(self, service, sample_products):
        result = service.generate_factory_order_excel(
            products=sample_products,
            boat_departure=date(2026, 1, 15),
        )
        wb = load_workbook(result)
        ws = wb.active

        # Check first product (row 8)
        assert ws["A8"].value == "ALMENDRO BEIGE"
        assert ws["A9"].value == "CEIBA GRIS CLARO"
        assert ws["A10"].value == "NOGAL CAFE"

    def test_excel_has_correct_m2(self, service, sample_products):
        result = service.generate_factory_order_excel(
            products=sample_products,
            boat_departure=date(2026, 1, 15),
        )
        wb = load_workbook(result)
        ws = wb.active

        # 14 pallets * 134.4 = 1881.6
        assert ws["C8"].value == 1881.6
        # 7 pallets * 134.4 = 940.8
        assert ws["C9"].value == 940.8
        # 14 pallets * 134.4 = 1881.6
        assert ws["C10"].value == 1881.6

    def test_excel_has_format_column(self, service, sample_products):
        result = service.generate_factory_order_excel(
            products=sample_products,
            boat_departure=date(2026, 1, 15),
        )
        wb = load_workbook(result)
        ws = wb.active

        assert ws["B8"].value == "51X51"
        assert ws["B9"].value == "51X51"
        assert ws["B10"].value == "51X51"

    def test_excel_has_total(self, service, sample_products):
        result = service.generate_factory_order_excel(
            products=sample_products,
            boat_departure=date(2026, 1, 15),
        )
        wb = load_workbook(result)
        ws = wb.active

        # Find TOTAL row (row 12 after 3 products + gap)
        assert ws["A12"].value == "TOTAL"
        # Total: 1881.6 + 940.8 + 1881.6 = 4704.0
        assert ws["C12"].value == 4704.0

    def test_excel_has_container_count(self, service, sample_products):
        result = service.generate_factory_order_excel(
            products=sample_products,
            boat_departure=date(2026, 1, 15),
        )
        wb = load_workbook(result)
        ws = wb.active

        # 35 pallets / 14 = 3 containers (ceil of 2.5)
        assert ws["A14"].value == "3 CONTENEDORES"

    def test_skips_zero_pallet_products(self, service):
        products = [
            {"sku": "ALMENDRO BEIGE BTE", "pallets": 14},
            {"sku": "SKIP THIS ONE", "pallets": 0},
            {"sku": "NOGAL CAFE BTE", "pallets": 7},
        ]
        result = service.generate_factory_order_excel(
            products=products,
            boat_departure=date(2026, 1, 15),
        )
        wb = load_workbook(result)
        ws = wb.active

        # Only 2 products should be in the list
        assert ws["A8"].value == "ALMENDRO BEIGE"
        assert ws["A9"].value == "NOGAL CAFE"
        # Row 10 is empty (gap), row 11 is TOTAL
        assert ws["A11"].value == "TOTAL"

    def test_empty_products_list(self, service):
        result = service.generate_factory_order_excel(
            products=[],
            boat_departure=date(2026, 1, 15),
        )
        wb = load_workbook(result)
        ws = wb.active

        # Should still have structure
        assert ws["A1"].value == "Pedido Tarragona Guatemala"
        assert ws["A9"].value == "TOTAL"
        assert ws["C9"].value == 0.0
        assert ws["A11"].value == "0 CONTENEDORES"


class TestGetExportService:
    """Tests for singleton pattern."""

    def test_returns_same_instance(self):
        service1 = get_export_service()
        service2 = get_export_service()
        assert service1 is service2

    def test_returns_export_service_instance(self):
        service = get_export_service()
        assert isinstance(service, ExportService)
